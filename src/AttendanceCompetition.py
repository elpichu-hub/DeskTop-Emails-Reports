import chardet
import pandas as pd
from fuzzywuzzy import process, fuzz
import openpyxl
from openpyxl.comments import Comment
import os

names_list = [
    "Amarillis Santiago",
    "Aniuska Rodriguez",
    "Ashley Kellyman",
    "Guadalupe Ocampo",
    "Kenelm Doleyres",
    "Mayelin Ferrales",
    "Sean Scott",
    "Tieheka Clarke",
    "Vanessa Oyola",
    "Waheeda Baksh"
]

def attendance_competition(file_path, competitionFile, weekly_attendance=None):
    with open(file_path, 'rb') as f:
        result = chardet.detect(f.read())  

    encoding = result['encoding']
    df = pd.read_csv(file_path, encoding=encoding)

    df['Date'] = pd.to_datetime(df['Date'])

    date_str = df['Date'].dt.strftime('%Y-%m-%d').iloc[0]

    # Load the workbook
    wb = openpyxl.load_workbook(competitionFile)

    # Select the active worksheet (or specify a sheet by name)
    ws = wb.active

    if weekly_attendance:
        # 1. Load the weekly_attendance into a DataFrame
        with open(weekly_attendance, 'rb') as f:
            result = chardet.detect(f.read())
        encoding = result['encoding']
        df_weekly = pd.read_csv(weekly_attendance, encoding=encoding)

        # 2. For each name in names_list, search for its best match using fuzzywuzzy
        filtered_rows = []

        for name in names_list:
            best_match, score, idx = process.extractOne(name, df_weekly['FullName'])
            if score >= 90:
                # Adding the entire row to the filtered list
                filtered_rows.append(df_weekly.iloc[idx])

        # 3. Create df_weekly_filtered from the filtered list
        df_weekly_filtered = pd.DataFrame(filtered_rows)
        print(df_weekly_filtered)

        filtered_names_set = set(df_weekly_filtered['FullName'])  # assuming 'FullName' is the column with names in df_weekly_filtered

        # Iterate through the rows in the worksheet to find each agent name, starting from row 21
        for row in ws.iter_rows(min_row=21, max_row=ws.max_row):  # updated min_row to 18
            agent_name = row[0].value  # assuming agent name is in column 1 (A)
            def is_name_match(agent_name, names_set, threshold=90):
                """
                Checks if agent_name has a match in names_set that exceeds the given threshold.
                """
                for name in names_set:
                    if fuzz.ratio(agent_name, name) >= threshold:
                        return True
                return False

            # Check if agent_name is in filtered_names_set
            if is_name_match(agent_name, filtered_names_set):
                # Get the current score from column D (assuming it's the score column)
                current_score = row[3].value  # assuming score is in column 4 (D)
                if current_score is None:
                    # If the score is None (perhaps because it's blank), treat it as 0
                    current_score = 0
                # Add 10 to the current score
                updated_score = current_score + 10
                # Update the cell value with the new score
                row[3].value = updated_score
                
                # Optional: Add a comment to indicate why 10 points were added
                comment_text = f"10 points added for weekly perfect attendance. {agent_name}\n\n"
                if row[3].comment:
                    # Append text to existing comment
                    row[3].comment.text += comment_text
                else:
                    # Create a new comment
                    comment = Comment(text=comment_text, author="Attendance Update Script")
                    # Assign the comment to the cell
                    row[3].comment = comment


    def rearrange_name(name):
        # If name is not a string, return it unchanged (or convert to string if desired)
        if not isinstance(name, str):
            return name  # or str(name) to convert to string, but be cautious with NaN values
        # Split the name on the comma
        parts = name.split(',')
        if len(parts) == 2:
            # If there are two parts, rearrange them
            return f"{parts[1].strip()} {parts[0].strip()}"
        # If there aren't two parts, return the name unchanged
        return name

    # Apply the rearrange_name function to each name in the 'Name' column
    df['Name'] = df['Name'].apply(rearrange_name)

    # Check for similar names using fuzzywuzzy
    matches = []

    for name in names_list:
        best_match, score, idx = process.extractOne(name, df['Name'])
        if score >= 90:
            call_out_type = df.loc[idx, 'Call Out Type']
            date_of_call_out = df.loc[idx, 'Date']
            reason = df.loc[idx, 'Reason']
            matches.append((name, best_match, score, call_out_type, date_of_call_out, reason))

    # Convert matches to a DataFrame for easier viewing
    matches_df = pd.DataFrame(matches, columns=['Original Name', 'Matched Name', 'Score', 'Call Out Type', 'date_of_call_out', 'reason'])

    # Convert matches_df to a dictionary for easy lookup
    matches_dict = dict(zip(matches_df['Original Name'], matches_df['Call Out Type']))
    
    # Iterate through the rows in the worksheet to find each agent name, starting from row 21
    for row in ws.iter_rows(min_row=21, max_row=ws.max_row):  # updated min_row to 18
        agent_name = row[0].value  # assuming agent name is in column 1 (A)
        if agent_name not in matches_dict:  # updated to check against keys in matches_dict
            # If the agent is not found in the matches_dict, add 5 to their attendance score
            current_attendance = row[1].value  # assuming attendance is in column 2 (B)
            if current_attendance is None:
                # If attendance is None (perhaps because it's blank), treat it as 0
                current_attendance = 0
            updated_attendance = current_attendance + 5
            row[1].value = updated_attendance

            # Create a comment with the date included
            comment_text = f"5 points added for not being on the call outs list on {date_str}. {agent_name}\n\n"
            
            # Check if cell already has a comment
            if row[1].comment:
                # Append text to existing comment
                row[1].comment.text += comment_text
            else:
                # Create a new comment
                comment = Comment(text=comment_text, author="Attendance Update Script")
                # Assign the comment to the cell
                row[1].comment = comment

        if agent_name in matches_dict:  # checking if agent_name is in the matches_dict
            # Get the Call Out Type for the agent
            call_out_type = matches_dict[agent_name]
            if call_out_type == "Absent":
                # If the Call Out Type is 'Absent', prepare the comment text
                # Assuming there's a column 'Date' in matches_df
                date_of_absence = matches_df.loc[matches_df['Original Name'] == agent_name, 'date_of_call_out'].values[0]
                date_of_absences = pd.to_datetime(date_of_absence).strftime('%m/%d/%y')
                comment_text = f"Appears on Call Out List from {date_of_absences} due to {call_out_type}. \n\n"
                
                # Check if cell already has a comment
                if row[1].comment:
                    # Append text to existing comment
                    row[1].comment.text += comment_text
                else:
                    # Create a new comment
                    comment = Comment(text=comment_text, author="Attendance Update Script")
                    # Assign the comment to the cell
                    row[1].comment = comment
            else:
                # Get the current score from column C (assuming it's the score column)
                current_score = row[2].value  # assuming score is in column 3 (C)
                if current_score is None:
                    # If the score is None (perhaps because it's blank), treat it as 0
                    current_score = 0
                # Subtract 2 from the current score
                updated_score = current_score - 2
                # Update the cell value with the new score
                row[2].value = updated_score

                date_of_absence = matches_df.loc[matches_df['Original Name'] == agent_name, 'date_of_call_out'].values[0]
                date_of_absences = pd.to_datetime(date_of_absence).strftime('%m/%d/%y')
                # Prepare the comment text
                comment_text = f"2 points subtracted for being on the call outs list on {date_of_absences} due to {call_out_type}. \n\n"
                
                # Check if cell already has a comment
                if row[2].comment:
                    # Append text to existing comment
                    row[2].comment.text += comment_text
                else:
                    # Create a new comment
                    comment = Comment(text=comment_text, author="Attendance Update Script")
                    # Assign the comment to the cell
                    row[2].comment = comment

    # Save the updated workbook to a new file (or overwrite the existing file if desired)
    wb.save(competitionFile)
    os.startfile(competitionFile)



