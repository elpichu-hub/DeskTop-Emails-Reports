from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from time import sleep
import os
from datetime import datetime, timedelta, date
import email_config
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException
import pandas as pd
from openpyxl import styles
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import logging


log_filename = "proponisiWinners.log"
# This function sends an email with an attachment
def send_email(subject, recipient, body, file_path=None):
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    import os
    from email.mime.base import MIMEBase
    from email import encoders
    

    EMAIL_ADDRESS = email_config.EMAIL_ADDRESS_AUTO
    EMAIL_PASSWORD = email_config.EMAIL_PASSWORD_AUTO

    # Create the email message
    message = MIMEMultipart()
    message['From'] = EMAIL_ADDRESS
    message['To'] = recipient
    message['Subject'] = subject

    message.attach(MIMEText(body, 'html'))

    if file_path is not None:
        with open(file_path, 'rb') as file:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(file.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', 'attachment', filename=os.path.basename(file_path))
            message.attach(part)
    

    # Connect to the Gmail SMTP server and send the email
    with smtplib.SMTP('smtp.gmail.com', 587) as smtp:
        smtp.starttls()
        smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
        smtp.send_message(message)

def get_dates_for_query():
    # Get today's date and time
    today = datetime.today()


    # Find the most recent Friday
    current_weekday = today.weekday()
    days_to_subtract = (current_weekday - 4) % 7
    if days_to_subtract == 0:
        days_to_subtract = 7
    last_friday = today - timedelta(days=days_to_subtract)

    # find the Saturday before the most recent Friday
    last_saturday = last_friday - timedelta(days=6)

    # Format the dates
    formatted_last_friday = last_friday.strftime("%m/%d/%Y")
    formatted_last_saturday = last_saturday.strftime("%m/%d/%Y")
    logging.info('Getting dates for query... Worked!')

    return formatted_last_saturday, formatted_last_friday


def check_teams(cleaned_table_data, driver, wait, actions):
    print('Starting check_teams function')
    driver.get("https://www.tpcdm.com")

    # Initialize the default team for each row in cleaned_table_data
    for data in cleaned_table_data:
        if len(data) < 4:  # If the team field is not already set
            data.append("MISSING_TEAM")

    for index, data in enumerate(cleaned_table_data):
        print(f'Processing data row {index}: {data}')
        try:
            sleep(5)
            my_team_icon = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "a.displayMyTeam i.fa-users")))
            driver.execute_script("arguments[0].click();", my_team_icon)
            sleep(5)
            search_box = wait.until(EC.element_to_be_clickable((By.ID, "search-teams")))
            sleep(2)

            sleep(2)
            search_box.send_keys(data[1])
            search_box.click()
            actions.send_keys(Keys.TAB).send_keys(Keys.TAB).send_keys(Keys.ENTER).perform()
            sleep(7)

            team_titles = [
                'Incredibles', 'Guardians', 'Jedi', 'Legacy', 'Marvel',
                'Gladiators', 'Elite', 'Padawan', 'Mandalorian', 'Hogwarts', 'Avatar'
                ]

            for team_title in team_titles:
                try:
                    element = wait.until(EC.visibility_of_element_located(
                        (By.CSS_SELECTOR, f"li.select2-selection__choice[title='{team_title}']")))
                    data[3] = team_title  # Update the team name
                    print(f"Found team: {team_title}")
                    break  # Exit the loop since we found the team
                except TimeoutException:
                    print(f"Team not found: {team_title}")
                    continue  # Continue to the next team_title

        except Exception as e:
            print("Error occurred during team checking:", e)

    print('check_teams function completed')
    logging.info('Checking teams... Worked!')
    print(cleaned_table_data)

    return cleaned_table_data



def create_format_file(last_saturday, last_friday, *table_data_args):
    # Create a new Excel workbook
    workbook = Workbook()

    # Create a new sheet
    sheet = workbook.active
    sheet.title = 'Combined Data'

    filename = 'proponisiWeeklyMontlyWinners.xlsx'

    for idx, table_data in enumerate(table_data_args, start=1):
        # Convert the table data to a DataFrame
        df = pd.DataFrame(table_data, columns=['Ranking', 'Name', 'Points', 'Team'])

        # Write headers for each table
        header_row = [f'{header}' for header in df.columns]
        sheet.append(header_row)

        # Convert DataFrame to a list of lists (convert NumPy arrays to lists)
        data_as_list = df.values.tolist()

        # Write data for each table
        for row_data in data_as_list:
            sheet.append(row_data)

    # -----start formatting-----

    # Set the width of the first four columns to 24 pixels
    for col_num in range(1, 5):
        col_letter = get_column_letter(col_num)
        sheet.column_dimensions[col_letter].width = 24
    
    # Insert a new row above row 13 and 7
    sheet.insert_rows(7)
    sheet.insert_rows(14)

    # Set the font for the first row
    blue = styles.PatternFill(start_color='0ACDFF', end_color='0ACDFF', fill_type='solid')
    red = styles.PatternFill(start_color='E6F9AF', end_color='E6F9AF', fill_type='solid')
    green = styles.PatternFill(start_color='59FFA0', end_color='59FFA0', fill_type='solid')
    orange = styles.PatternFill(start_color='FF7F11', end_color='FF7F11', fill_type='solid')
    first_row = sheet[1]
    for cell in first_row:
        cell.fill = red

    nineth_row = sheet[8]
    for cell in nineth_row:
        cell.fill = red

    seventeenth_row = sheet[15]
    for cell in seventeenth_row:
        cell.fill = red

    sheet.insert_rows(1)
    sheet.insert_rows(8)
    sheet.insert_rows(17)

    # Merge cells A1 to D1
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    # Set the merged cell's value (e.g., a title)
    sheet.cell(row=1, column=1, value="Merged Title")


    # Merge cells A9 to D9
    sheet.merge_cells(start_row=9, start_column=1, end_row=9, end_column=4)

    # Set the merged cell's value (e.g., a title)
    sheet.cell(row=9, column=1, value="Merged Title for Row 9")

    # Merge cells A9 to D9
    sheet.merge_cells(start_row=17, start_column=1, end_row=17, end_column=4)

    # Set the merged cell's value (e.g., a title)
    sheet.cell(row=17, column=1, value="Merged Title for Row 9")

    # add color to 1st row
    first_row = sheet[1]
    for cell in first_row:
        cell.fill = blue

    # add color to 9th row
    nineth_row = sheet[9]
    for cell in nineth_row:
        cell.fill = green

    # add color to 17th row
    seventeenth_row = sheet[17]
    for cell in seventeenth_row:
        cell.fill = orange

    # format for full document
    # Set the height for all rows with data to 20
    for row in sheet.iter_rows(min_row=1):  # Start from the second row (skip headers)
        for cell in row:
            sheet.row_dimensions[cell.row].height = 20

    # Center-align and center text in the entire worksheet
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Apply borders and font styling to rows with data
    for row in sheet.iter_rows(min_row=1):  # Start from the second row (skip headers)
        for cell in row:
            if cell.value is not None:
                border = Border(left=Side(style='medium'), 
                                right=Side(style='medium'),
                                top=Side(style='medium'),
                                bottom=Side(style='medium'))
                cell.border = border
                cell.font = Font(bold=True)

    # after merge
     # Apply borders to the merged cell range A1:D1
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            border = Border(left=Side(style='medium'), 
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))
            cell.border = border

    # Apply borders to the merged cell range A9:D9
    for row in sheet.iter_rows(min_row=9, max_row=9):
        for cell in row:
            border = Border(left=Side(style='medium'), 
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))
            cell.border = border

    # Apply borders to the merged cell range A17:D17
    for row in sheet.iter_rows(min_row=17, max_row=17):
        for cell in row:
            border = Border(left=Side(style='medium'), 
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium'))
            cell.border = border

    sheet['A1'] = f'Call Center: {last_saturday} - {last_friday}'
    sheet['A9'] = f'Escalations/Web Chat: {last_saturday} - {last_friday}'
    sheet['A17'] = f'Correspondence Clerks: {last_saturday} - {last_friday}'

    
    try:
        # Specify the directory and filename
        directory = "Gamification_winners"
        formatted_date = f'{last_saturday} - {last_friday}'.replace('/', '_')
        filename = os.path.join(directory, f'proponisiWeeklyMontlyWinners{formatted_date}.xlsx')

        # Create the directory if it doesn't exist
        if not os.path.exists(directory):
            os.makedirs(directory)

        # Save the workbook with the specified filename
        workbook.save(filename)
    except Exception as e:
        print(f"An error occurred: {e}")

    logging.info('Creating and formatting file... Worked!')

    return filename

def clean_names(data):
    """
    Takes a list of lists, where each inner list contains a name at index 1.
    Cleans up the names by removing any extra spaces or non-standard whitespace characters.

    :param data: List of lists, where each inner list contains a name at index 1.
    :return: List of lists with cleaned names.
    """
    cleaned_data = []
    for row in data:
        if len(row) > 1 and isinstance(row[1], str):
            # Normalize space characters and strip leading/trailing whitespace
            name = ' '.join(row[1].split())
            # Replace the original name with the cleaned name
            row[1] = name
            cleaned_data.append(row)
    return cleaned_data


# This function navigates to the login page, logs in, sets the dates and downloads the file
def navigate_login_get_data(formatted_last_saturday, formatted_last_friday, level, driver, wait, actions):

    # Maximize the window
    driver.maximize_window()

    # Navigate to the URL
    driver.get("https://www.tpcdm.com/Account/Login?ReturnUrl=%2f")

    # Find and fill the form fields based on the data from the current row
    username_field = driver.find_element("id", "Email")
    username_field.send_keys('lazaro.gonzalez@conduent.com')
    username_field.send_keys(Keys.ENTER)

    password_field = driver.find_element("id", "Password")
    password_field.send_keys('Pinkfloyd(07)!')
    password_field.send_keys(Keys.ENTER)

    print('Logged In!')

    driver.get("https://www.tpcdm.com/Gaming/Game/GameRanking")
    
    # Use WebDriverWait to wait for the element to be present in the DOM and to be clickable.
    view_all_radio = wait.until(EC.element_to_be_clickable((By.ID, "Options")))
    # Once the element is ready, use JavaScript to click on it.
    driver.execute_script("arguments[0].click();", view_all_radio)

    
    try:
        # Wait and click the combobox to open it
        group_field = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@role="combobox"]')))
        group_field.click()
        print('clicked')


        # Wait for the dropdown to be interactable after click
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".select2-search__field")))

        # Move to the search box within the dropdown, simulate key down and enter
        if level == 1:
            actions.send_keys(Keys.ARROW_DOWN).send_keys(Keys.ENTER).perform()
        if level == 2:
            actions.send_keys(Keys.ARROW_DOWN).send_keys(Keys.ARROW_DOWN).send_keys(Keys.ENTER).perform()
        if level == 3:
            actions.send_keys(Keys.ARROW_DOWN).send_keys(Keys.ARROW_DOWN).send_keys(Keys.ARROW_DOWN).send_keys(Keys.ENTER).perform()

        print(f'level {level} selected')

    except Exception as e:
        print(e)

    try:
        # Interact with the start date field
        start_date_field = driver.find_element(By.ID, "StartDate")
        start_date_field.click()
        start_date_field.clear()
        print(formatted_last_saturday)
        start_date_field.send_keys(formatted_last_saturday)

        # Interact with the end date field
        end_date_field = driver.find_element(By.ID, "EndDate")
        end_date_field.click()  # Clicking the field before clearing might be necessary
        for _ in range(20):  # Send backspace 20 times to clear the field
            end_date_field.send_keys(Keys.BACK_SPACE)
        end_date_field.send_keys(formatted_last_friday)

    except Exception as e:
        print(e)

    submit_button = driver.find_element("id", "btnSubmit")  
    submit_button.click()

    sleep(10)  # Adjust the sleep duration if needed

    # Initialize a list to hold all row data
    table_data = []
    
    try:
        # Now find all the rows within the table body
        rows = driver.find_elements(By.XPATH, "//table/tbody/tr")[:5]

        # Iterate over each row
        for row in rows:
            # Find all the cells within the row
            cells = row.find_elements(By.XPATH, ".//td")  # The dot at the beginning ensures that it only searches within the context of the row

            # Retrieve the text from each cell and add it to a row_data list
            row_data = [cell.text for cell in cells]
            
            # Add the row_data list to the table_data list
            table_data.append(row_data)

    except Exception as e:
        print(e)

    # CLEAN THE DATA
    cleaned_table_data = [[cell for cell in row if cell] for row in table_data]
    print(cleaned_table_data)

    # Clean the names
    cleaned_table_data = clean_names(cleaned_table_data)
    


    sleep(5)

    logging.info('Navigating, logging in, setting dates and getting data... Worked!')

    # return cleaned_table_data
    return cleaned_table_data


def main_func():
    try:
        # Set the download directory path
        download_dir = r'C:\MyStuff\Projects\Developing\Tests\StatsOrginizer\src\downloads_proponisi'
        # Set the options
        chrome_options = webdriver.ChromeOptions()
        # set preferences for the download directory
        prefs = {
            "download.default_directory": download_dir,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True
        }
        # Add the preferences to the options
        chrome_options.add_experimental_option("prefs", prefs)
        # Specify the path to the chromedriver.exe executable (modify as needed)
        chrome_driver_path = r'C:\MyStuff\Projects\Developing\scraper\chromedriver-win64\chromedriver.exe'
        # Set the path to the chromedriver executable as an environment variable
        os.environ["webdriver.chrome.driver"] = chrome_driver_path
        # Initialize the WebDriver without specifying executable_path
        driver = webdriver.Chrome(options=chrome_options)
        # Initialize the WebDriverWait with a timeout of 10 seconds
        wait = WebDriverWait(driver, 10)

        # Create an instance of ActionChains
        actions = ActionChains(driver)

        # ------------start calling functions ------------

        # # Get the dates for the query
        last_saturday, last_friday = get_dates_for_query()

        # Navigate to the login page, log in, set the dates and get the data for all levels
        cleaned_table_data1 = navigate_login_get_data(last_saturday, last_friday, 1, driver, wait, actions)        
        cleaned_table_data2 = navigate_login_get_data(last_saturday, last_friday, 2, driver, wait, actions)
        cleaned_table_data3 = navigate_login_get_data(last_saturday, last_friday, 3, driver, wait, actions)

        # Check the teams
        try:
            cleaned_table_data_team1 = check_teams(cleaned_table_data1, driver, wait, actions)  
            cleaned_table_data_team2 = check_teams(cleaned_table_data2, driver, wait, actions)  
            cleaned_table_data_team3 = check_teams(cleaned_table_data3, driver, wait, actions)  
        except Exception as e:
            print(e)

        # Close the WebDriver after processing all rows
        driver.quit()

        # Create and format the Excel file
        file_to_send = create_format_file(last_saturday, last_friday, cleaned_table_data_team1, cleaned_table_data_team2, cleaned_table_data_team3)

        body = f"""
        Good morning, Please see below winners of Gamification for the week of {last_saturday} - {last_friday}. In addition, please remember the first prize for each week earns an extra 15 minutes of SBK. They will be provided with a pass that they will need to show to their respective team lead/supervisor to be able to claim the prize. Please let me know if you have any additional questions.
        """

        # Send an email with the file as an attachment
        send_email(f"Proponisi Winners {last_saturday} - {last_friday}", 'lazaro.gonzalez@conduent.com', body, file_to_send)

        # Log the end of the script
        logging.info('Script execution completed!')
    except Exception as e:
        print(e)
        logging.error('Failed to run script.', exc_info=True)
    

main_func()
    

    