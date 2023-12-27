import pandas as pd
import chardet
import datetime
import random
import os
import openpyxl
import email_config
import os

def send_email(subject, recipient, body, file_path, sbs_form_mentors, cc=None, bcc=None):
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    import os
    from email.mime.base import MIMEBase
    from email import encoders
    

    EMAIL_ADDRESS = email_config.EMAIL_ADDRESS_AUTO
    EMAIL_PASSWORD = email_config.EMAIL_PASSWORD_AUTO

    # Create the email message
    message = MIMEMultipart()
    message['From'] = EMAIL_ADDRESS
    message['To'] = recipient
    message['Subject'] = subject

    message.attach(MIMEText(body, 'html'))

    if file_path is not None:
        with open(file_path, 'rb') as file:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(file.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', 'attachment', filename=os.path.basename(file_path))
            message.attach(part)
    
    if sbs_form_mentors is not None:
        with open(sbs_form_mentors, 'rb') as mentors_file:
            mentors_part = MIMEBase('application', 'octet-stream')
            mentors_part.set_payload(mentors_file.read())
            encoders.encode_base64(mentors_part)
            mentors_part.add_header('Content-Disposition', 'attachment', filename=os.path.basename(sbs_form_mentors))
            message.attach(mentors_part)

    # Connect to the Gmail SMTP server and send the email
    with smtplib.SMTP('smtp.gmail.com', 587) as smtp:
        smtp.starttls()
        smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
        smtp.send_message(message)

def detect_encoding(file_path):
    with open(file_path, 'rb') as f:
        result = chardet.detect(f.read())
    return result['encoding']


def create_sbs(file_path, email_address, team, throughEmail=True):

    encoding = detect_encoding(file_path)
    df = pd.read_csv(file_path, encoding=encoding)

    now = datetime.datetime.now() + datetime.timedelta(minutes=30)
    # now = datetime.datetime.now().replace(hour=10, minute=30, second=0, microsecond=0)

    df_with_schedules = df.dropna(subset=['SCHEDULE START', 'SCHEDULE END'])

    mentors = df_with_schedules[df_with_schedules['Type'] == 'Mentor']
    mentees = df[df['Type'] == 'Mentee']

    mentors = mentors.sample(frac=1, random_state=42).reset_index(drop=True)

    sbs_pairs = pd.DataFrame(columns=['Mentor', 'Mentee', 'Time', 'Team'])
    
    max_end_time = datetime.datetime.combine(datetime.date.today(), datetime.time(15, 30))

    occupied_time_slots = []

    for _, mentor in mentors.iterrows():
        mentor_name = mentor['NAME']
        mentor_bilingual = mentor['Bilingual']
        mentor_lunch_start = pd.to_datetime(mentor['LUNCH START'], format='%I:%M %p').time() if pd.notna(mentor['LUNCH START']) else None
        mentor_lunch_end = pd.to_datetime(mentor['LUNCH END'], format='%I:%M %p').time() if pd.notna(mentor['LUNCH END']) else None
        mentor_schedule_end = pd.to_datetime(mentor['SCHEDULE END'], format='%I:%M %p').time()

        # If the mentor has a lunch break, add it to 'occupied_time_slots'
        if mentor_lunch_start is not None and mentor_lunch_end is not None:
            lunch_start_datetime = datetime.datetime.combine(datetime.date.today(), mentor_lunch_start)
            lunch_end_datetime = datetime.datetime.combine(datetime.date.today(), mentor_lunch_end)
            occupied_time_slots.append((lunch_start_datetime, lunch_end_datetime))

        eligible_mentees = mentees.loc[(mentees['Bilingual'] == mentor_bilingual) | (mentees['Bilingual'].isnull())]

        if not eligible_mentees.empty:
            mentee = eligible_mentees.sample(n=1)
            mentee_name = mentee.iloc[0]['NAME']
            mentee_lunch_start = pd.to_datetime(mentee.iloc[0]['LUNCH START'], format='%I:%M %p').time() if pd.notna(mentee.iloc[0]['LUNCH START']) else None
            mentee_lunch_end = pd.to_datetime(mentee.iloc[0]['LUNCH END'], format='%I:%M %p').time() if pd.notna(mentee.iloc[0]['LUNCH END']) else None

            mentor_schedule_start = pd.to_datetime(mentor['SCHEDULE START'], format='%I:%M %p').time()

            end_of_intervals = datetime.datetime.combine(datetime.date.today(), mentor_schedule_end) - datetime.timedelta(minutes=45)
            time_intervals = pd.date_range(start=datetime.datetime.combine(datetime.date.today(), mentor_schedule_start),
                                        end=end_of_intervals,
                                        freq='15min')

            valid_time_intervals = [time for time in time_intervals if time >= now and time + datetime.timedelta(minutes=45) <= max_end_time]

            # Considering mentor's and mentee's lunch break
            if mentor_lunch_start is not None and mentor_lunch_end is not None or mentee_lunch_start is not None and mentee_lunch_end is not None:
                lunch_start_datetime_mentor = datetime.datetime.combine(datetime.date.today(), mentor_lunch_start) if mentor_lunch_start is not None else None
                lunch_end_datetime_mentor = datetime.datetime.combine(datetime.date.today(), mentor_lunch_end) if mentor_lunch_end is not None else None

                lunch_start_datetime_mentee = datetime.datetime.combine(datetime.date.today(), mentee_lunch_start) if mentee_lunch_start is not None else None
                lunch_end_datetime_mentee = datetime.datetime.combine(datetime.date.today(), mentee_lunch_end) if mentee_lunch_end is not None else None

                valid_time_intervals = [time for time in valid_time_intervals if not (
                    (lunch_start_datetime_mentor is not None and lunch_end_datetime_mentor is not None and ((lunch_start_datetime_mentor < time < lunch_end_datetime_mentor) or (lunch_start_datetime_mentor < time + datetime.timedelta(minutes=45) <= lunch_end_datetime_mentor))) or
                    (lunch_start_datetime_mentee is not None and lunch_end_datetime_mentee is not None and ((lunch_start_datetime_mentee <= time < lunch_end_datetime_mentee) or (lunch_start_datetime_mentee < time + datetime.timedelta(minutes=45) < lunch_end_datetime_mentee)))
                )]

            if len(valid_time_intervals) > 0:
                random.shuffle(valid_time_intervals)

                for start_time in valid_time_intervals:
                    end_time = start_time + datetime.timedelta(minutes=45)


                    if any([(occupied_start <= start_time < occupied_end) or 
                            (occupied_start < end_time <= occupied_end) or 
                            (start_time <= occupied_start < end_time) or 
                            (start_time < occupied_start and end_time > occupied_end) for occupied_start, occupied_end in occupied_time_slots]):
                        continue


                    # If the SBS session does not overlap with the mentor's lunch break or an already scheduled SBS session, then schedule it
                    start_time_str = start_time.strftime('%I:%M %p')
                    end_time_str = end_time.strftime('%I:%M %p')

                    time_str = f"{start_time_str} - {end_time_str}"

                    sbs_pairs = pd.concat([sbs_pairs, pd.DataFrame({'Mentor': [mentor_name], 'Mentee': [mentee_name], 'Time': [time_str]})], ignore_index=True)
                    mentees = mentees.drop(mentee.index)
                    occupied_time_slots.append((start_time, end_time))

                    break


    # Second loop to include remaining mentees
    for _, mentee in mentees.iterrows():
        mentee_name = mentee['NAME']
        mentee_bilingual = mentee['Bilingual']

        # Filter mentors based on the bilingual status of the mentee and if they have not been scheduled already
        eligible_mentors = mentors[~mentors['NAME'].isin(sbs_pairs['Mentor'])]
        eligible_mentors = eligible_mentors if pd.isnull(mentee_bilingual) else eligible_mentors[eligible_mentors['Bilingual'] == 'Y']

        # If there are no eligible mentors left, add mentee to sbs_pairs without mentor and time
        if eligible_mentors.empty:
            sbs_pairs = pd.concat([sbs_pairs, pd.DataFrame({'Mentor': [None], 'Mentee': [mentee_name], 'Time': [None]})],
                                ignore_index=True)
            continue

        # Choose a random mentor for each remaining mentee
        mentor = eligible_mentors.sample(n=1)
        mentor_name = mentor.iloc[0]['NAME']
        mentor_lunch_start = pd.to_datetime(mentor.iloc[0]['LUNCH START'], format='%I:%M %p').time() if pd.notna(mentor.iloc[0]['LUNCH START']) else None
        mentor_lunch_end = pd.to_datetime(mentor.iloc[0]['LUNCH END'], format='%I:%M %p').time() if pd.notna(mentor.iloc[0]['LUNCH END']) else None
        mentor_schedule_start = pd.to_datetime(mentor.iloc[0]['SCHEDULE START'], format='%I:%M %p').time()

        # To this
        end_of_intervals = datetime.datetime.combine(datetime.date.today(), mentor_schedule_end) - datetime.timedelta(minutes=45)
        time_intervals = pd.date_range(start=datetime.datetime.combine(datetime.date.today(), mentor_schedule_start),
                                    end=end_of_intervals,
                                    freq='15min')

        valid_time_intervals = [time for time in time_intervals if time >= now and time + datetime.timedelta(minutes=45) <= max_end_time]

       # Considering mentor's and mentee's lunch break
        if mentor_lunch_start is not None and mentor_lunch_end is not None or mentee_lunch_start is not None and mentee_lunch_end is not None:
            lunch_start_datetime_mentor = datetime.datetime.combine(datetime.date.today(), mentor_lunch_start) if mentor_lunch_start is not None else None
            lunch_end_datetime_mentor = datetime.datetime.combine(datetime.date.today(), mentor_lunch_end) if mentor_lunch_end is not None else None

            lunch_start_datetime_mentee = datetime.datetime.combine(datetime.date.today(), mentee_lunch_start) if mentee_lunch_start is not None else None
            lunch_end_datetime_mentee = datetime.datetime.combine(datetime.date.today(), mentee_lunch_end) if mentee_lunch_end is not None else None

            valid_time_intervals = [time for time in valid_time_intervals if not (
                (lunch_start_datetime_mentor is not None and lunch_end_datetime_mentor is not None and ((lunch_start_datetime_mentor < time < lunch_end_datetime_mentor) or (lunch_start_datetime_mentor < time + datetime.timedelta(minutes=45) <= lunch_end_datetime_mentor))) or
                (lunch_start_datetime_mentee is not None and lunch_end_datetime_mentee is not None and ((lunch_start_datetime_mentee <= time < lunch_end_datetime_mentee) or (lunch_start_datetime_mentee < time + datetime.timedelta(minutes=45) < lunch_end_datetime_mentee)))
            )]


        if len(valid_time_intervals) > 0:
            start_time = random.choice(valid_time_intervals)
            end_time = start_time + datetime.timedelta(minutes=45)

            start_time_str = start_time.strftime('%I:%M %p')
            end_time_str = end_time.strftime('%I:%M %p')

            time_str = f"{start_time_str} - {end_time_str}"

            sbs_pairs = pd.concat([sbs_pairs, pd.DataFrame({'Mentor': [mentor_name], 'Mentee': [mentee_name], 'Time': [time_str]})],
                                ignore_index=True)

    sbs_pairs['Team'] = team.capitalize()



    team_leads = {
        'Elite': 'William / Karina / Ernesto',
        'Gladiators': 'Ura / Daisy / Iqbal',
        'Marvel': 'Omaira / Johanna / Kathleen',
        'Legacy': 'Yari / Shaynika / Edgar',
        'Avatar': 'Lissette / Jose / Mara / Krystal / Natacha',
        'Hogwarts': 'Farzanna / Jordan / Brittnay / Luis',
        'Bootcamp': 'Melanie / Daniqua / Edgar / Jose / Jordan / Brittany',
        'N/A': 'N/A'
    }

    sbs_pairs['Sup / Leads'] = team_leads[team.capitalize()]

    today = datetime.date.today().strftime('%m.%d.%y')

    # Create the Excel file path
    file_path = f'Side By Side {today}.xlsx'

    # Save the SBS pairs to the Excel file
    sbs_pairs = sbs_pairs.sort_values('Time')  # Sort the DataFrame by the 'Time' column

    # Save the SBS pairs to the Excel file
    sbs_pairs.to_excel(file_path, index=False)

    print(sbs_pairs)

    # Open the Excel file using openpyxl
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Add borders to all cells with data and color the background of unscheduled mentees
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin')
            )
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Add yellow background for bilingual mentees
            if cell.column == 2:  # Assuming 'Mentee' column is the second column (column index 2)
                mentee_name = cell.value
                mentee = df[df['NAME'] == mentee_name]  # Use the original df to get the mentee information
                if not mentee.empty:
                    mentee = mentee.iloc[0]
                    orange = openpyxl.styles.PatternFill(start_color='FF7F11', end_color='FF7F11', fill_type='solid')
                    if mentee['Bilingual'] == 'Y':
                        cell.fill = orange

            # If mentee couldn't be scheduled, color the background of the row red
            red = openpyxl.styles.PatternFill(start_color='C1292E', end_color='C1292E', fill_type='solid')
            if cell.row > 1 and ws.cell(row=cell.row, column=1).value is None:
                cell.fill = red

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 42

    # Set the row height for rows A1, B1, and C1
    ws.row_dimensions[1].height = 45

    # Set the blue background color for rows A1, B1, and C1
    blue = openpyxl.styles.PatternFill(start_color='0ACDFF', end_color='0ACDFF', fill_type='solid')
    for col in range(1, 6):
        cell = ws.cell(row=1, column=col)
        cell.fill = blue
        cell.font = openpyxl.styles.Font(size=16)

    # Set the font size to 12 for rows 2 and below
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = openpyxl.styles.Font(size=12)


    # Save the modified Excel file
    wb.save(file_path)
    print(file_path)
    os.startfile(file_path)

    if throughEmail == True:


        # Determine the appropriate greeting based on the current time
        if now.hour < 12:
            greeting = "Good morning"
        elif now.hour < 18:
            greeting = "Good afternoon"
        else:
            greeting = "Good evening"

        print(f"Excel file '{file_path}' created successfully with borders added!")

        subject = f'Side By Side {today} {team.capitalize()}'
        # body = "<html>\n" \
        #    "<body>\n" \
        #    f"<h3>{greeting} everyone,</h3>\n" \
        #    "<h3>Please find below your scheduled Side By Side (SBS) sessions.</h3>\n" \
        #    "<h3>Please advise your agents of the SBS.</h3>\n" \
        #    "<h3>The lines in red are cancelled due to the agents not being at work at the scheduled time or day or due to availability</h3>\n" \
        #    "<h3>Please have a splitter cord available for the SBS.</h3>\n" \
        #    "<h3>If your agent is not able to participate in the SBS, please let us know before we send out our agent.</h3>\n" \
        #    f"<h3>Mentors, please complete the enclosed SBS form and share it with the supervisor and team leads of the agents, {team_leads[team]}.</h3>\n" \
        #    f"<br>{email_config.team_lead_elite_signature}\n" \
        #    "</body>\n" \
        #    "</html>"
        
        if team != 'Bootcamp':
            body = "<html>\n" \
        "<body>\n" \
        f"<h3>{greeting} everyone,</h3>\n" \
        "<h3>Please find below your scheduled Side By Side (SBS) sessions.</h3>\n" \
        "<h3>Kindly modify your schedules to accommodate mentor support, allowing you to receive assistance as needed.<h3>\n" \
        "<h3>Please have a splitter cord available for the SBS.</h3>\n" \
        "<h3>If your agent is not able to participate in the SBS, please let us know before we send out our agent.</h3>\n" \
        f"<h3>Mentors, please complete the enclosed SBS form and share it with the supervisor and team leads of the agents, including {team_leads[team]}.</h3>\n" \
        f"<br>{email_config.team_lead_elite_signature}\n" \
        "</body>\n" \
        "</html>"
        else:
            body = "<html>\n" \
        "<body>\n" \
        f"<h3>{greeting} everyone,</h3>\n" \
        "<h3>Please find below your scheduled Side By Side (SBS) sessions.</h3>\n" \
        "<h3>Kindly modify your schedules to accommodate mentor support, allowing you to receive assistance as needed.<h3>\n" \
        "<h3>Please have a splitter cord available for the SBS.</h3>\n" \
        "<h3>If your agent is not able to participate in the SBS, please let us know before we send out our agent.</h3>\n" \
        f"<h3>Mentors, kindly fill out the attached SBS form and forward it to the Bootcamp Management Team, represented by {team_leads[team]}.</h3>\n" \
        f"<br>{email_config.team_lead_elite_signature}\n" \
        "</body>\n" \
        "</html>"
        
        sbs_form_mentors = 'SBS Mentor Form.docx'

        send_email(subject, email_address, body, file_path, sbs_form_mentors)

# file_path = 'mentors_mentees.csv'
# email_address = 'lazarogonzalez.auto@gmail.com'
# team = 'Elite'
# create_sbs(file_path, email_address, team)