def send_email(subject, recipient, body, img_path_meme=None, img_path_Work=None, stats_data=None, cc=None, bcc=None):
    import email_config
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    from email.mime.image import MIMEImage
    import os
    from email.mime.base import MIMEBase
    from email import encoders
    import encouraging_and_success_messages
    

    EMAIL_ADDRESS = email_config.EMAIL_ADDRESS_AUTO
    EMAIL_PASSWORD = email_config.EMAIL_PASSWORD_AUTO

    # Create the email message
    message = MIMEMultipart()
    message['From'] = EMAIL_ADDRESS
    message['To'] = recipient
    message['Subject'] = subject
    message['Cc'] = cc
    message['Bcc'] = bcc

    message.attach(MIMEText(body, 'html'))

    # If an image path is provided, add the image as an inline attachment
    if img_path_meme is not None:
        with open(img_path_meme, 'rb') as img_file:
            img_data = img_file.read()
        img_mime = MIMEImage(img_data)
        img_mime.add_header('Content-ID', '<{}>'.format(os.path.basename(img_path_meme)))
        img_mime.add_header('Content-Disposition', 'inline', filename=os.path.basename(img_path_meme))
        message.attach(img_mime)

    # If an img_path_100 is provided, add the image as an inline attachment
    if img_path_Work is not None:
        with open(img_path_Work, 'rb') as img_file:
            img_data = img_file.read()
        img_mime = MIMEImage(img_data)
        img_mime.add_header('Content-ID', '<{}>'.format(os.path.basename(img_path_Work)))
        img_mime.add_header('Content-Disposition', 'inline', filename=os.path.basename(img_path_Work))
        message.attach(img_mime)

    # If a stats_data file is provided, add it as an attachment
    if stats_data is not None:
        with open(stats_data, 'rb') as file:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(file.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', 'attachment', filename=os.path.basename(stats_data))
            message.attach(part)

    # Connect to the Gmail SMTP server and send the email
    with smtplib.SMTP('smtp.gmail.com', 587) as smtp:
        smtp.starttls()
        smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
        smtp.send_message(message)

# file_path = 'User Productivity Summary DAILY.xls'
# email_address = 'lazarogonzalez.auto@gmail.com'
# run daily stats Non csv file just excel
def run_daily_stats(file_path, email_address, cc=None, bcc=None, team=None):
    from openpyxl import load_workbook
    from openpyxl import utils, styles
    import os
    from datetime import datetime, time
    import pandas as pd
    import os
    import random
    import email_config
    import encouraging_and_success_messages

    # this will make sure that no matter which extentions
    # the file is saved the program still runs
    # xls or xlsx
    try:
        df = pd.read_excel(file_path)
        file_name, file_ext = os.path.splitext(os.path.basename(file_path))
        file_name = file_name.replace('/', '-')
        new_file_path = os.path.join(os.path.dirname(file_path), file_name + '.xlsx')
        df.to_excel(new_file_path, index=False)
    except Exception as e:
        print(f'An error occurred: {e}')

    # This variable will save the stat type
    # to use in A1 and N29
    stats_type = ''
        
    # with an excel workbook created load it to 
    # update it. The file needs to have extension .xlsx,
    # .xls won't work
    # if the file does not exists this will return
    # a value that will be printed on the app text area.
    try:
        work_book = load_workbook(new_file_path)
    except Exception as e:
        print(e)
        return e

    # this selects the main sheet in the workbook
    work_sheet = work_book.active

    # Get the report Date from file without changes
    # and format the date to xx/xx/xxxx
    report_date_from_ICBM = work_sheet['F5'].value
    def find_row_for_phrase(sheet, phrase):
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and phrase in str(cell.value):
                    return cell.row
        return None

    # Your phrase to search
    search_phrase = 'Date:'

    # Find the row containing the search_phrase
    date_row = find_row_for_phrase(work_sheet, search_phrase)

    # Check if the row was found
    if date_row is not None:
        # Get the cell from column F (column number 6) on the same row
        date_cell = work_sheet.cell(row=date_row, column=6)

        # Extract the date value
        report_date_from_ICBM = date_cell.value
    else:
        print(f'Phrase "{search_phrase}" not found in the worksheet')


    # this will separate the two dates from ICBM
    # to determine the amount of days. 
    # it will get start date and end date as a string
    date_range = report_date_from_ICBM.split(" - ")
    start_date_str = date_range[0].strip()
    end_date_str = date_range[1].split(" ")[0].strip()

    # convert the date strings to datetime objects
    start_date = datetime.strptime(start_date_str, "%m/%d/%Y %I:%M:%S %p")
    end_date = datetime.strptime(end_date_str, "%m/%d/%Y")
    delta = end_date - start_date
    # number of days to determine if the stats are
    # daily, weekly or monthly
    num_days = delta.days

    # Format the dates for the document titles if weekly,
    # monthly or yearly
    start_date_formatted = start_date.date().strftime("%m.%d.%y")
    end_date_formatted = end_date.date().strftime("%m.%d.%y")

    # Extract the start date from the string
    report_date_string = report_date_from_ICBM.split(" - ")[0]
    start_date = datetime.strptime(report_date_string, "%m/%d/%Y %I:%M:%S %p")
    day_of_the_week = start_date.strftime('%A')

    # Format the start date as a string with only the date
    report_date = start_date.strftime("%m/%d/%Y")

    # This will loop through all the images
    # on the workbook and delete them all.
    for image in work_sheet._images:
        # Remove the image
        work_sheet._images.remove(image)

    # work_sheet.delete_rows(1, 11)
    # Find the row index where the phrase "Summaries Per User" is located
    def find_row_for_phrase(sheet, phrase):
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and phrase in str(cell.value):
                    return cell.row
        return None

    # Your phrase to search
    search_phrase = 'Summaries  Per User'

    # Find the row containing the search_phrase
    row_to_delete_until = find_row_for_phrase(work_sheet, search_phrase)

    # Check if the row was found
    if row_to_delete_until is not None:
        # Delete the rows from the start to the found row (inclusive)
        work_sheet.delete_rows(1, row_to_delete_until)
    else:
        print(f'Phrase "{search_phrase}" not found in the worksheet')
    
    work_sheet.delete_rows(2)
    work_sheet.delete_rows(4, 2)

    # set a starting points to start seaching
    # empty rows after such point
    start_row = 3

    # Loop through the rows in the worksheet and delete all empty rows
    for row in range(start_row, work_sheet.max_row + 1):
        if all(cell.value is None for cell in work_sheet[row]):
            empty_row = row
            print(f'Next Empty Row {empty_row}')
            work_sheet.delete_rows(empty_row, 1)

    # Set the target phrase to search for
    target_phrase = 'Summaries Per User And Queue'

    # Find the row that contains the target phrase
    for row in work_sheet.iter_rows(min_row=1, max_col=work_sheet.max_column):
        for cell in row:
            if cell.value == target_phrase:
                # Delete the row that contains the target phrase
                work_sheet.delete_rows(cell.row, 1)
                # Delete all rows below the target phrase
                work_sheet.delete_rows(cell.row, work_sheet.max_row - cell.row + 1)
                break



    # At his point all the important data is all remaining 
    # from the rows. Columns Cleaning starts here.
    # ---------------------------------------------------------------------- #

    # Set the width of all columns to 10 
    for column in range(1, work_sheet.max_column + 1):
        column_letter = utils.get_column_letter(column)
        work_sheet.column_dimensions[column_letter].width = 12

    # Delete cols:
    work_sheet.delete_cols(2, 4)
    work_sheet.delete_cols(4, 6)
    work_sheet.delete_cols(5, 3)
    work_sheet.delete_cols(11, 3)


    # Delete all columns after handle time or 11
    num_cols = work_sheet.max_column - 11
    # Delete the columns
    if num_cols > 0:
        work_sheet.delete_cols(12, num_cols)

    # At this points only the necessary rows and columns
    # are displayed
    # ------------------------------------------------------------------------- #  

    # will add in a1 daily, weekly, monthly or yearly
    # depending on the amount of days been calculated
    if num_days == 0:
        work_sheet['A1'].value = f'Daily Stats {day_of_the_week} {report_date}'
        stats_type = 'Daily'
    if num_days > 1 and num_days < 20:
        work_sheet['A1'].value = f'Weekly Stats {start_date_formatted} - {end_date_formatted}'
        stats_type = 'Weekly'
    if num_days > 20 and num_days < 300:
        work_sheet['A1'].value = f'Monthly Stats {start_date_formatted} - {end_date_formatted}'
        stats_type = 'Monthly'
    if num_days > 300:
        work_sheet['A1'].value = f'Yearly Stats {start_date_formatted} - {end_date_formatted}'


    print(num_days)

    work_sheet['B1'].value = 'Offered'
    work_sheet['B2'].value = '#'
    work_sheet['K1'].value = 'Handle Time'
    work_sheet['K2'].value = 'Average'
    
    if team in email_config.web_chat_variations:
        work_sheet['E1'].value = 'Chat Time'
    else:
        work_sheet['E1'].value = 'Talk Time'

    # styling
    # Define the border style
    thin_border = styles.Border(
        left=styles.Side(style='thin'),
        right=styles.Side(style='thin'),
        top=styles.Side(style='thin'),
        bottom=styles.Side(style='thin')
    )




    # add width to specific columns
    work_sheet.column_dimensions['A'].width = 37
    work_sheet.column_dimensions['D'].width = 14
    work_sheet.column_dimensions['K'].width = 15

    # merge cells
    work_sheet.merge_cells('A1:A2')
    work_sheet.merge_cells('E1:F1')
    work_sheet.merge_cells('G1:H1')
    work_sheet.merge_cells('I1:J1')

    merged_range = 'A1:A2'

    # Create a Style object with centered alignment
    center_alignment = styles.Alignment(horizontal='center', vertical='center')

    # Loop over all rows and columns and apply the centered alignment
    for row in work_sheet.iter_rows():
        for cell in row:
            cell.alignment = center_alignment

    # Set the height of rows 1 and 2 to 20 and all other rows to 17
    for row in range(1, work_sheet.max_row + 1):
        if row == 1 or row == 2:
            work_sheet.row_dimensions[row].height = 20
            work_sheet.row_dimensions[row].font = styles.Font(size=14, bold=True)
        else:
            work_sheet.row_dimensions[row].height = 17
            work_sheet.row_dimensions[row].font = styles.Font(size=11)


    # Set the height of rows 1 and 2 to 20 and all other rows to 17
    for row in range(1, work_sheet.max_row + 1):
        if row == 1 or row == 2:
            work_sheet.row_dimensions[row].height = 25
            for cell in work_sheet[row]:
                cell.font = styles.Font(size=12, bold=True)
        else:
            work_sheet.row_dimensions[row].height = 17
            for cell in work_sheet[row]:
                cell.font = styles.Font(size=11)


    # this will loop thorugh all the rows from 4 to the last one
    # only rows with data will be looped through 
    new_data = []
    for row in range(4, work_sheet.max_row + 1):
        values = [cell.value for cell in work_sheet[row]]
        if all(value is None for value in values):
            continue  # skip empty rows
        if 'User' in str(values):  # skip rows that contain 'User'
            continue
        new_data.append(values)

    new_data = [row for row in new_data if all(val is not None for val in row)]

    # # this will sort the data based on the names
    # print("Sorting data based on names...")
    # sorted_data = sorted(new_data, key=lambda x: x[0] or '') 

    # this will sort the data based on the AHT
    print("Sorting data based on AHT...")
    sorted_data = sorted(new_data, key=lambda x: x[10] or '') 

    # this gets the data sorted out based on the handle time average
    print("Sorting data based on handle time average...")
    sorted_data_for_top_5 = sorted(new_data, key=lambda x: x[10] or '')
    
    # this are the top 5
    print("Getting top 5...")
    top_5 = sorted_data_for_top_5[:5]
    print(top_5)

    # delete the data in the work_sheet and added 
    # sorted out
    for row in work_sheet.iter_rows(min_row=4):
        for cell in row:
            cell.value = None

    # I don't understand this lines quite well,
    # this will assigned all the sorted data
    # to each row and cell.
    for i, row in enumerate(sorted_data, start=4):
        for j, value in enumerate(row):
            work_sheet.cell(row=i, column=j+1, value=value)

    # this will make the 3rd row grey.
    grey = styles.PatternFill(start_color='808080', end_color='808080', fill_type='solid')
    for row in work_sheet.iter_rows(min_row=3, max_row=3):
        for cell in row:
            cell.fill = grey

    green = styles.PatternFill(start_color='59FFA0', end_color='59FFA0', fill_type='solid')
    orange = styles.PatternFill(start_color='FF7F11', end_color='FF7F11', fill_type='solid')
    red = styles.PatternFill(start_color='C1292E', end_color='C1292E', fill_type='solid')

    if team == None:
        # on column K color based on conditions
        for cell in work_sheet['K']:
            try:
                time_obj = datetime.strptime(str(cell.value), "%H:%M:%S").time()
                # if 6:30 or less green
                if isinstance(time_obj, time) and time_obj < time(hour=0, minute=6, second=30):
                    cell.fill = green
                if isinstance(time_obj, time) and time_obj <= time(hour=0, minute=7, second=0) and time_obj > time(hour=0, minute=6, second=30):
                    cell.fill = orange
                if isinstance(time_obj, time) and time_obj > time(hour=0, minute=7, second=0):
                    cell.fill = red
            except ValueError:
                pass

        # more than 30 seconds will be red
        for cell in work_sheet['J']:
            try:
                time_obj = datetime.strptime(str(cell.value), "%H:%M:%S").time()
                # if 6:30 or less green
                if isinstance(time_obj, time) and time_obj > time(hour=0, minute=0, second=20):
                    cell.fill = red
                else:
                    cell.fill = green
            except ValueError:
                pass


        # Get the cell J3 and its value
        j3_cell = work_sheet['J3']
        value = j3_cell.value

        # if more than 30
        time_obj = datetime.strptime(str(j3_cell.value), "%H:%M:%S").time()
        if isinstance(time_obj, time) and time_obj > time(hour=0, minute=0, second=20):
            j3_cell.fill = red
            print(f'red ${j3_cell.value}')
        else:
            print('green')
            j3_cell.fill = green


        # more than 30 seconds will be red
        for cell in work_sheet['H']:
            try:
                time_obj = datetime.strptime(str(cell.value), "%H:%M:%S").time()
                # if 6:30 or less green
                if isinstance(time_obj, time) and time_obj > time(hour=0, minute=0, second=30):
                    cell.fill = red
            except ValueError:
                pass

        # more than 7 minutes will be red
        for cell in work_sheet['F']:
            try:
                time_obj = datetime.strptime(str(cell.value), "%H:%M:%S").time()
                # if 6:30 or less green
                if isinstance(time_obj, time) and time_obj > time(hour=0, minute=7, second=0):
                    cell.fill = red
            except ValueError:
                pass

    elif team in email_config.web_chat_variations:
        # on column K color based on conditions
        for cell in work_sheet['K']:
            try:
                time_obj = datetime.strptime(str(cell.value), "%H:%M:%S").time()
                # if 6:30 or less green
                if isinstance(time_obj, time) and time_obj <= time(hour=0, minute=7, second=30):
                    cell.fill = green
                if isinstance(time_obj, time) and time_obj <= time(hour=0, minute=8, second=0) and time_obj > time(hour=0, minute=7, second=30):
                    cell.fill = orange
                if isinstance(time_obj, time) and time_obj > time(hour=0, minute=8, second=0):
                    cell.fill = red
            except ValueError:
                pass

        # more than 30 seconds will be red
        for cell in work_sheet['J']:
            try:
                time_obj = datetime.strptime(str(cell.value), "%H:%M:%S").time()
                # if 6:30 or less green
                if isinstance(time_obj, time) and time_obj > time(hour=0, minute=0, second=20):
                    cell.fill = red
                else:
                    cell.fill = green
            except ValueError:
                pass


        # Get the cell J3 and its value
        j3_cell = work_sheet['J3']
        value = j3_cell.value

        # if more than 30
        time_obj = datetime.strptime(str(j3_cell.value), "%H:%M:%S").time()
        if isinstance(time_obj, time) and time_obj > time(hour=0, minute=0, second=30):
            j3_cell.fill = red
        else:
            print('green')
            j3_cell.fill = green


        # more than 30 seconds will be red
        for cell in work_sheet['H']:
            try:
                time_obj = datetime.strptime(str(cell.value), "%H:%M:%S").time()
                # if 6:30 or less green
                if isinstance(time_obj, time) and time_obj > time(hour=0, minute=0, second=30):
                    cell.fill = red
            except ValueError:
                pass

        # more than 7 minutes will be red
        for cell in work_sheet['F']:
            try:
                time_obj = datetime.strptime(str(cell.value), "%H:%M:%S").time()
                # if 6:30 or less green
                if isinstance(time_obj, time) and time_obj > time(hour=0, minute=8, second=0):
                    cell.fill = red
            except ValueError:
                pass

    elif team in email_config.escalations_variations:
        # on column K color based on conditions
        for cell in work_sheet['K']:
            try:
                time_obj = datetime.strptime(str(cell.value), "%H:%M:%S").time()
                # if 6:30 or less green
                if isinstance(time_obj, time) and time_obj <= time(hour=0, minute=10, second=30):
                    cell.fill = green
                if isinstance(time_obj, time) and time_obj <= time(hour=0, minute=11, second=0) and time_obj > time(hour=0, minute=10, second=30):
                    cell.fill = orange
                if isinstance(time_obj, time) and time_obj > time(hour=0, minute=11, second=0):
                    cell.fill = red
            except ValueError:
                pass

        # more than 30 seconds will be red
        for cell in work_sheet['J']:
            try:
                time_obj = datetime.strptime(str(cell.value), "%H:%M:%S").time()
                # if 6:30 or less green
                if isinstance(time_obj, time) and time_obj > time(hour=0, minute=0, second=20):
                    cell.fill = red
                else:
                    cell.fill = green
            except ValueError:
                pass


        # Get the cell J3 and its value
        j3_cell = work_sheet['J3']
        value = j3_cell.value

        # if more than 30
        time_obj = datetime.strptime(str(j3_cell.value), "%H:%M:%S").time()
        if isinstance(time_obj, time) and time_obj > time(hour=0, minute=0, second=30):
            j3_cell.fill = red
        else:
            print('green')
            j3_cell.fill = green


        # more than 30 seconds will be red
        for cell in work_sheet['H']:
            try:
                time_obj = datetime.strptime(str(cell.value), "%H:%M:%S").time()
                # if 6:30 or less green
                if isinstance(time_obj, time) and time_obj > time(hour=0, minute=0, second=30):
                    cell.fill = red
            except ValueError:
                pass

        # more than 7 minutes will be red
        for cell in work_sheet['F']:
            try:
                time_obj = datetime.strptime(str(cell.value), "%H:%M:%S").time()
                # if 6:30 or less green
                if isinstance(time_obj, time) and time_obj > time(hour=0, minute=11, second=0):
                    cell.fill = red
            except ValueError:
                pass


    blue = styles.PatternFill(start_color='0ACDFF', end_color='0ACDFF', fill_type='solid')
    # make row 1 and 2 blue
    for row in work_sheet.iter_rows(min_row=1, max_row=2):
        for cell in row:
            cell.fill = blue



    # Apply the border style to cells with data
    for row in work_sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.border = thin_border


    # this is the list of all the merged cells ranges
    # I put them here so I can add border with the 
    # functions add_borders_to_merged_cells
    merged_ranges = ['A1:A2', 'E1:F1', 'G1:H1', 'I1:J1', 'N29:R30', 'N31:O32']

    
    # creates a date string to name the file
    if stats_type == 'Daily':
        date_for_file_name = f'{start_date.strftime("%m.%d.%Y")}'
    else:
        date_for_file_name = f'{start_date_formatted}-{end_date_formatted}'



    # Find the index of column K
    k_col_index = None
    for col in range(1, work_sheet.max_column + 1):
        if work_sheet.cell(row=1, column=col).value == 'Handle Time':
            k_col_index = col
            break

    # Delete rows that do not have a value in column K
    rows_to_delete = []
    for row in range(4, work_sheet.max_row + 1):
        if work_sheet.cell(row=row, column=k_col_index).value is None:
            rows_to_delete.append(row)

    # Remove rows in reverse order to avoid changing the index of rows to delete
    for row in reversed(rows_to_delete):
        work_sheet.delete_rows(row)


    # top 5 table creation starts here
    # ---------------------------------------- #
    work_sheet.merge_cells('N29:R30')
    work_sheet['N29'].value = 'Top 5!'
    work_sheet['N29'].alignment  = center_alignment
    work_sheet['N29'].font = styles.Font(size=14, bold=True)
    work_sheet['N29'].fill = blue
    work_sheet.merge_cells('N31:O32')

    if stats_type == 'Daily':
        work_sheet['N31'].value = f'Daily Stats {day_of_the_week} {report_date}'
    if stats_type == 'Weekly':
        work_sheet['N31'].value = f'Weekly Stats {start_date_formatted} - {end_date_formatted}'
    if stats_type == 'Monthly':
        work_sheet['N31'].value = f'Montly Stats {start_date_formatted} - {end_date_formatted}'
    if stats_type == 'Yearly':
        work_sheet['N31'].value = f'Yearly Stats {start_date_formatted} - {end_date_formatted}'

    work_sheet['N31'].alignment  = center_alignment 
    work_sheet['N31'].font = styles.Font(size=12, bold=True)
    work_sheet.column_dimensions['N'].width = 24.71
    work_sheet['P31'].value = 'Offered'
    work_sheet['P32'].value = '#'
    work_sheet['Q31'].value = 'Answered'
    work_sheet['Q32'].value = '#'
    work_sheet['R31'].value = 'Handle Time'
    work_sheet.column_dimensions['R'].width = 15
    work_sheet['R32'].value = 'Average'
    work_sheet['A3'].value = encouraging_and_success_messages.motivating_short_phrases[random.randint(0, len(encouraging_and_success_messages.motivating_short_phrases) - 1)]

    work_sheet.merge_cells('N33:O33')
    work_sheet.merge_cells('N34:O34')
    work_sheet.merge_cells('N35:O35')
    work_sheet.merge_cells('N36:O36')
    work_sheet.merge_cells('N37:O37')

    work_sheet['A3'].fill = green

    # loop through each row in the range N33 to R37
    for row in work_sheet['N33:R37']:
        # loop through each cell in the row and apply the border
        for cell in row:
            cell.border = thin_border

     # this will itarate through the top 5 and 
    # assign the values to the cells
    # above is a very repetive way to do the same
     # Write top 5 employees to the worksheet
    for i, row in enumerate(top_5[:5], start=33):
        work_sheet[f'N{i}'].value = row[0]
        work_sheet[f'P{i}'].value = row[1]
        work_sheet[f'Q{i}'].value = row[2]
        work_sheet[f'R{i}'].value = row[10]
        work_sheet[f'R{i}'].fill = green
        # center and align the content in the cell
        for col in ('N', 'P', 'Q', 'R'):
            cell = work_sheet[f'{col}{i}']
            cell.alignment = styles.Alignment(horizontal='center', vertical='center')
    
    for row in range(31, 33):
        print(row)
        for col in ['N', 'O', 'P', 'Q', 'R']:
            cell = work_sheet[f'{col}{row}']
            cell.alignment = center_alignment
            cell.fill = blue
            cell.font = styles.Font(size=12, bold=True)

    # this will add borders to the ranges from merged_ranges
    def add_borders_to_merged_cells(merged_ranges):
        for merged_range in merged_ranges:
            for row in work_sheet[merged_range]:
                for cell in row:
                    cell.border = thin_border


    add_borders_to_merged_cells(merged_ranges)

    # loop through each row in the range N33 to R37
    for row in work_sheet['N29:R37']:
        # loop through each cell in the row and apply the border
        for cell in row:
            cell.border = thin_border


    # define arguments for send_email
    subject = None
    if stats_type == 'Daily':
        if team in email_config.web_chat_variations:
            subject = f'{stats_type} Stats {day_of_the_week} {date_for_file_name} WebChat'
        elif team in email_config.escalations_variations:
            subject = f'{stats_type} Stats {day_of_the_week} {date_for_file_name} Escalations'   
        else:
            subject = f'{stats_type} Stats {day_of_the_week} {date_for_file_name}'   
    else:
        if team in email_config.web_chat_variations:
            subject = f'{stats_type} Stats {start_date_formatted} - {end_date_formatted} WebChat'
        elif team in email_config.escalations_variations:
            subject = f'{stats_type} Stats {start_date_formatted} - {end_date_formatted} Escalations'
        else:
            subject = f'{stats_type} Stats {start_date_formatted} - {end_date_formatted}'  
    
    

    # Get current working directory
    current_dir = os.getcwd()

    # Specify the folders with the images
    meme_folder = os.path.join(current_dir, 'memes_images')
    work_folder = os.path.join(current_dir, 'work_images')

    # Check that the folders exist
    if not os.path.exists(meme_folder) or not os.path.exists(work_folder):
        raise Exception("One or both image folders do not exist")

    # Get the current day of the week
    current_day = datetime.now().strftime('%A').lower()

    # Get the meme file for the current day
    meme_file = f"{current_day}.jpg"  # replace '.jpg' with your image extension

    # Check if the meme file exists
    if not os.path.isfile(os.path.join(meme_folder, meme_file)):
        raise Exception(f"Meme for {current_day} not found")

    # Select a random file from the work files
    work_files = [f for f in os.listdir(work_folder) if os.path.isfile(os.path.join(work_folder, f))]
    random_work_file = random.choice(work_files)

    img_path_meme = os.path.join(meme_folder, meme_file)
    img_path_Work = os.path.join(work_folder, random_work_file)

    # Get the current date and time
    now = datetime.now()
    greeting = ""

    # Determine the appropriate greeting based on the current time
    if now.hour < 12:
        greeting = "Good morning"
    elif now.hour < 18:
        greeting = "Good afternoon"
    else:
        greeting = "Good evening"
    
    # Add HTML tags
    body = "<html><body>"

    # Add result string and signature
    result_string = f"<h3>{greeting} Team, \n\nPlease find below your {subject}. <span style='background-color: #59FFA0';>{encouraging_and_success_messages.motivational_phrases[random.randint(0, len(encouraging_and_success_messages.motivational_phrases) - 1)]}</span></h3> "
    body += result_string.replace('\n', '<br>')

    # Include the meme image
    body += "<img src='cid:{}' alt='Meme Image' style='width: 519px; height: 305px;'><br>".format(os.path.basename(img_path_meme))

    # Include the work image
    body += "<img src='cid:{}' alt='Work Image' style='width: 519px; height: 600px;'><br>".format(os.path.basename(img_path_Work))

    # Add signature to the body
    body += email_config.team_lead_elite_signature

    # Close HTML tags
    body += "</body></html>"

    # this will save the changes. 
    # The work book needs to be closed 
    # in order to save changes. 
    # if the stats are daily save it with 
    # name as the date of the report
    # if weekly, montly or yearly then
    # save with the range of the report
    # if an error ocurrs it will send it
    # to statsOrginizer to show on the app.
    # here it would be that the file is opened
    # at the moment PermissionError
    if stats_type == 'Daily':
        try:
            file_to_open = f'{stats_type}_report_lazaro_gonzalez_{date_for_file_name}.xlsx'
            work_book.save(file_to_open)
            # os.startfile(file_to_open)
        except Exception as e:
            print(e)
            return e
    else:
        try:
            file_to_open = f'{stats_type}_report_lazaro_gonzalez_{start_date_formatted}-{end_date_formatted}.xlsx'
            work_book.save(file_to_open)
            # os.startfile(file_to_open)
        except Exception as e:
            print(e)
            return e

    
    # Send the email for the current user
    send_email(subject, email_address, body, img_path_meme=img_path_meme, img_path_Work=img_path_Work, stats_data=file_to_open ,cc=cc, bcc=bcc)

    
