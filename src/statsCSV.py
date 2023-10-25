    # Create a function to send the email
def send_email(subject, recipient, body, img_path=None, img_path_100=None):
    import email_config
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    from email.mime.image import MIMEImage
    import os

    EMAIL_ADDRESS = email_config.EMAIL_ADDRESS
    EMAIL_PASSWORD = email_config.EMAIL_PASSWORD

    # Create the email message
    message = MIMEMultipart()
    message['From'] = EMAIL_ADDRESS
    message['To'] = recipient
    message['Subject'] = subject
    message.attach(MIMEText(body, 'html'))

    # If an image path is provided, add the image as an inline attachment
    if img_path is not None:
        with open(img_path, 'rb') as img_file:
            img_data = img_file.read()
        img_mime = MIMEImage(img_data)
        img_mime.add_header('Content-ID', '<{}>'.format(os.path.basename(img_path)))
        img_mime.add_header('Content-Disposition', 'inline', filename=os.path.basename(img_path))
        message.attach(img_mime)

    # If an img_path_100 is provided, add the image as an inline attachment
    if img_path_100 is not None:
        with open(img_path_100, 'rb') as img_file:
            img_data = img_file.read()
        img_mime = MIMEImage(img_data)
        img_mime.add_header('Content-ID', '<{}>'.format(os.path.basename(img_path_100)))
        img_mime.add_header('Content-Disposition', 'inline', filename=os.path.basename(img_path_100))
        message.attach(img_mime)

    # Connect to the Gmail SMTP server and send the email
    with smtplib.SMTP('smtp.gmail.com', 587) as smtp:
        smtp.starttls()
        smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
        smtp.send_message(message)

def seconds_to_hms(seconds):
    h = int(seconds // 3600)
    m = int((seconds % 3600) // 60)
    s = int(seconds % 60)
    return f"{h:02d}:{m:02d}:{s:02d}"

# format the date entered if it has / in it
def format_date(date_string):
        return date_string.replace('/', '.')

# Function to convert time string in HH:MM:SS format to seconds
def time_to_seconds(time_str):
    h, m, s = time_str.split(':')
    return int(h) * 3600 + int(m) * 60 + int(s)

# Function to convert seconds to HH:MM:SS format
def seconds_to_time(seconds):
    if isinstance(seconds, str):
        return seconds
    m, s = divmod(int(seconds), 60)
    h, m = divmod(m, 60)
    return f"{h:02d}:{m:02d}:{s:02d}"


date_string = '5.27.2023 - 6.02.2023'
file_path1 = 'User Productivity Summary stats two call center 5.15.2023 teams.csv'
def run_report_stats(date_string, file_path1):
    import os
    import pandas as pd
    import openpyxl
    from datetime import datetime, time
    import calendar
    from openpyxl import utils, styles
    from openpyxl.styles import PatternFill
    from encouraging_and_success_messages import motivating_short_phrases
    import random

    date_string = format_date(date_string)
    
    print(date_string, file_path1)
    # read the CSV file into a pandas dataframe
    df = pd.read_csv(file_path1)   

    # specify the column names you want to keep
    user_id = 'i3user'
    acd_calls = 'totEnteredACD' # calls offered
    acd_answered = 'totnAnsweredACD' # calls answered
    total_hold = 'totTHoldACD'
    total_acw = 'totTACW'
    tot_talk = 'totTTalkACD'
    full_name = 'DisplayUserName'
    calls_transfered = 'totnTransferedACD'
    team = 'JobTitle'
    
    # keep only selected columns in the dataframe
    df_custom = df[[user_id, acd_calls, acd_answered, total_hold, total_acw, tot_talk, full_name, calls_transfered, team]]

    # Get the unique team values
    unique_teams = df_custom[team].unique()

    # Specify the path to the Excel file
    output_path = f"{date_string} stats_by_lazaro.xlsx"

    # create a ExcelWriter object
    writer = pd.ExcelWriter(output_path, engine='openpyxl')
    # Iterate over the unique teams
    for team_name in unique_teams:
        # Filter the DataFrame to keep only rows for the current team
        team_df = df_custom[df_custom[team] == team_name]

        # group the data by the 'i3user' column and calculate the total 'totEnteredACD',   'totTHoldACD', 'totTACW', and 'totTTalkACD'
        df_grouped = team_df.groupby(user_id).agg({
            acd_calls: 'sum',
            acd_answered: 'sum',
            total_hold: 'sum',
            total_acw: 'sum',
            tot_talk: 'sum',
            calls_transfered: 'sum',
            full_name: 'first',
        }).reset_index()
        
        
        if df_grouped.isna().any().any():
            df_grouped = df_grouped.fillna(0)
            print("Warning: The grouped data contains NaN values.")

        
        epsilon = 1e-8  # Small constant to avoid division by zero

        df_grouped['average_hold'] = (df_grouped[total_hold] / (df_grouped[acd_answered] + epsilon))
        df_grouped['total_hold'] = df_grouped[total_hold]

        df_grouped['average_acw'] = (df_grouped[total_acw] / (df_grouped[acd_answered] + epsilon))
        df_grouped['total_acw'] = df_grouped[total_acw]

        df_grouped['AHT'] = ((df_grouped[tot_talk] + df_grouped[total_acw]) / (df_grouped[acd_answered] + epsilon))

        df_grouped['calls_offered'] = df_grouped[acd_calls]
        df_grouped['calls_answered'] = df_grouped[acd_answered]
        df_grouped['calls_transfered'] = df_grouped[calls_transfered]

        df_grouped['total_talk'] = df_grouped[tot_talk]
        df_grouped['average_talk'] = (df_grouped[tot_talk] / df_grouped[acd_answered] + epsilon)


        # Remove rows where ACD Calls is 0
        df_grouped = df_grouped[df_grouped['calls_answered'] != 0]

        # Remove extra spaces in FullName column
        df_grouped['FullName'] = df_grouped['DisplayUserName'].str.strip().str.replace(r'\s+', ' ', regex=True)

        # Sort the dataframe based on the 'FullName' column (A to Z)
        df_grouped = df_grouped.sort_values(by='FullName')

        
        # # Reorder the columns as required
        columns = ['FullName', 'calls_offered', 'calls_answered', 'calls_transfered', 'total_talk', 'average_talk', 'total_hold', 'average_hold', 'total_acw', 'average_acw', 'AHT']
        df_grouped = df_grouped[columns]


        # Apply the helper function to the relevant columns
        for column in ['total_talk', 'average_talk', 'total_hold', 'average_hold', 'total_acw', 'average_acw', 'AHT']:
            df_grouped[column] = df_grouped[column].apply(seconds_to_time )

        # Reorder the columns as required
        columns = ['FullName', 'calls_offered', 'calls_answered', 'calls_transfered', 'total_talk', 'average_talk', 'total_hold', 'average_hold', 'total_acw', 'average_acw', 'AHT']
        df_grouped = df_grouped[columns]

        # get the totals of all the values to add to daily stats
        total_calls_offered = df_grouped['calls_offered'].sum()
        total_calls_answered = df_grouped['calls_answered'].sum()
        total_calls_transfered = df_grouped['calls_transfered'].sum()

        import math






       # Calculate the totals and format them back to time strings
        total_total_talk = seconds_to_time(math.ceil(df_grouped['total_talk'].apply(time_to_seconds).sum()))
        total_total_hold = seconds_to_time(math.ceil(df_grouped['total_hold'].apply(time_to_seconds).sum()))
        total_total_acw = seconds_to_time(math.ceil(df_grouped['total_acw'].apply(time_to_seconds).sum()))

        # Calculate the totals and format them back to time strings
        total_total_talk = seconds_to_time (total_total_talk)
        total_total_hold = seconds_to_time (total_total_hold)
        total_total_acw = seconds_to_time (total_total_acw)

       # Calculate the averages and format them back to time strings
        average_average_talk = seconds_to_time(math.ceil(df_grouped['average_talk'].apply(time_to_seconds).mean()))
        average_average_hold = seconds_to_time(math.ceil(df_grouped['average_hold'].apply(time_to_seconds).mean()))
        average_average_acw = seconds_to_time(math.ceil(df_grouped['average_acw'].apply(time_to_seconds).mean()))
        average_AHT = seconds_to_time(math.ceil(df_grouped['AHT'].apply(time_to_seconds).mean()))








        # Save the DataFrame to an Excel sheet for the current team
        df_grouped.to_excel(writer, sheet_name=team_name, index=False)

        # Sort the dataframe based on the 'AHT' column and get the top 5 rows
        top_5 = df_grouped.sort_values(by='AHT', ascending=True).head(5)

        wb = writer.book
        sheet = wb[team_name]

        # Iterate over all cells in the sheet and clear formatting
        for row in sheet.iter_rows():
            for cell in row:
                # Clear formatting from the cell
                cell.number_format = 'General'

        # Set the width of all columns to 10 
        for column in range(1, sheet.max_column + 1):
            column_letter = utils.get_column_letter(column)
            sheet.column_dimensions[column_letter].width = 15

        # Insert a row at the top and at row 3
        sheet.insert_rows(2)
        sheet.insert_rows(0)
        # Adjust the width of column A
        sheet.column_dimensions['A'].width = 49
        # Merge and center cells A1 and A2
        sheet.merge_cells('A1:A2')
        merged_cell = sheet['A1']
        merged_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')



        # Split the date range string into start and end dates
        date_range = date_string.split(' - ')  

        if len(date_range) == 1:  # Single date (daily stats)
            start_date_string = end_date_string = date_range[0]
            start_date = end_date = datetime.strptime(date_range[0], '%m.%d.%Y')
        else:  # Date range
            start_date_string = date_range[0]
            end_date_string = date_range[1].split(' ')[0]
            start_date = datetime.strptime(start_date_string, '%m.%d.%Y')
            end_date = datetime.strptime(end_date_string, '%m.%d.%Y')

        # Now that start_date_string has been defined, you can use it to get the day of the week:
        date_obj = datetime.strptime(start_date_string, '%m.%d.%Y')
        day_of_week = calendar.day_name[date_obj.weekday()]
        print(f'Day of the week {day_of_week}')

        delta = end_date - start_date
        num_days = delta.days + 1  # Adding 1 to include the end date

        stats_type = None
        # Determine the stats_type based on the number of days
        if num_days == 1:
            stats_type = 'Daily'
        elif num_days <= 7:
            stats_type = 'Weekly'
        elif num_days < 30:
            stats_type = 'Monthly'
        else:
            stats_type = 'Yearly'

        # Add value to cell A1
        if stats_type == 'Daily':
            merged_cell.value = f"{team_name} {stats_type} Stats {day_of_week} {date_string.replace('.', '/')}"
        else:
            merged_cell.value = f"{team_name} {stats_type} Stats {date_string.replace('.', '/')}"

        # Center all cells in the document
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        # Colors for background
        blue = PatternFill(start_color='0ACDFF', end_color='0ACDFF', fill_type='solid')
        green = PatternFill(start_color='59FFA0', end_color='59FFA0', fill_type='solid')
        orange = PatternFill(start_color='FF7F11', end_color='FF7F11', fill_type='solid')
        red = PatternFill(start_color='C1292E', end_color='C1292E', fill_type='solid')
        grey = PatternFill(start_color='808080', end_color='808080', fill_type='solid')

        # Set the values for cells B1, C1, and D1
        sheet['B1'].value = "Offered"
        sheet['C1'].value = "Answered"
        sheet['D1'].value = "Transferred"

        sheet['E1'].value = "Talk Time"
        sheet['G1'].value = "Hold Time"
        sheet['I1'].value = "ACW Time"
        sheet['K1'].value = "Handle Time"

        sheet['B2'].value = "#"
        sheet['C2'].value = "#"
        sheet['D2'].value = "#"
        sheet['E2'].value = "Duration"
        sheet['F2'].value = "Average"
        sheet['G2'].value = "Duration"
        sheet['H2'].value = "Average"
        sheet['I2'].value = "Duration"
        sheet['J2'].value = "Average"
        sheet['K2'].value = "Average"

        sheet['A3'].value = motivating_short_phrases[random.randint(0, len(motivating_short_phrases) - 1)]
        sheet['B3'].value = total_calls_offered
        sheet['C3'].value = total_calls_answered
        sheet['D3'].value = total_calls_transfered
        sheet['E3'].value = total_total_talk
        sheet['F3'].value = seconds_to_time (average_average_talk)
        sheet['G3'].value = total_total_hold
        sheet['H3'].value = seconds_to_time (average_average_hold)
        sheet['I3'].value = total_total_acw
        sheet['J3'].value = seconds_to_time (average_average_acw)
        sheet['K3'].value = seconds_to_time (average_AHT)
        

        # Merge cells E1 and F1
        sheet.merge_cells('E1:F1')

        # Merge cells G1 and H1
        sheet.merge_cells('G1:H1')

        # Merge cells I1 and J1
        sheet.merge_cells('I1:J1')

        # Get cell A1
        cell_a1 = sheet['A1']
        # Set the font style to bold
        cell_a1.font = styles.Font(bold=True)
        
        # Set the background color of rows 1 and 2 from column A to column K
        for row in sheet.iter_rows(min_row=1, max_row=2):
            for cell in row:
                cell.fill = blue
                cell.font = styles.Font(bold=True)

        # Set the background color of row 3 to grey
        for cell in sheet[3]:
            cell.fill = grey

        sheet['A3'].fill = green

        # Iterate over all cells in the worksheet
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = styles.Border(
                    left=styles.Side(style='thin'),
                    right=styles.Side(style='thin'),
                    top=styles.Side(style='thin'),
                    bottom=styles.Side(style='thin')
                )


        # Set the height of rows 1 and 2 to 25 and all other rows to 17
        for row in range(1, sheet.max_row + 1):
            if row == 1 or row == 2:
                sheet.row_dimensions[row].height = 25
                for cell in sheet[row]:
                    cell.font = styles.Font(size=12, bold=True)  # explicitly setting the font size for rows 1 and 2 to 12
            else:
                sheet.row_dimensions[row].height = 17
                for cell in sheet[row]:
                    cell.font = styles.Font(size=12)

        # on column K color based on conditions
        for cell in sheet['K']:
            try:
                time_obj = datetime.strptime(str(cell.value), "%H:%M:%S").time()
                if isinstance(time_obj, time) and time_obj <= time(hour=0, minute=6, second=30):
                    cell.fill = green
                elif isinstance(time_obj, time) and time_obj >= time(hour=0, minute=6, second=31) and time_obj <= time(hour=0, minute=7, second=0):
                    cell.fill = orange
                else:
                    cell.fill = red
            except ValueError:
                pass

        # more than 20 seconds will be red
        for cell in sheet['J']:
            try:
                time_obj = datetime.strptime(str(cell.value), "%H:%M:%S").time()
                if isinstance(time_obj, time) and time_obj > time(hour=0, minute=0, second=20):
                    cell.fill = red
                else:
                    cell.fill = green
            except ValueError:
                pass

        # Get the cell J3 and its value
        j3_cell = sheet['J3']
        value = j3_cell.value

        # if more than 30
        time_obj = datetime.strptime(str(j3_cell.value), "%H:%M:%S").time()
        if isinstance(time_obj, time) and time_obj > time(hour=0, minute=0, second=30):
            j3_cell.fill = red
        else:
            j3_cell.fill = green

        # more than 30 seconds will be red
        for cell in sheet['H']:
            try:
                time_obj = datetime.strptime(str(cell.value), "%H:%M:%S").time()
                if isinstance(time_obj, time) and time_obj > time(hour=0, minute=0, second=30):
                    cell.fill = red
            except ValueError:
                pass

        # more than 7 minutes will be red
        for cell in sheet['F']:
            try:
                time_obj = datetime.strptime(str(cell.value), "%H:%M:%S").time()
                if isinstance(time_obj, time) and time_obj > time(hour=0, minute=7, second=0):
                    cell.fill = red
            except ValueError:
                pass





        # top 5 area stats here
        # Create a Style object with centered alignment
        center_alignment = styles.Alignment(horizontal='center', vertical='center')

        sheet.merge_cells('N29:R30')
        sheet['N29'].value = 'Top 5!'
        sheet['N29'].alignment  = center_alignment
        sheet['N29'].font = styles.Font(size=14, bold=True)
        sheet['N29'].fill = blue
        sheet.merge_cells('N31:O32')

        
        if stats_type == 'Daily':
            sheet['N31'].value = f'Daily Stats {day_of_week} {date_string}'
        if stats_type == 'Weekly':
            sheet['N31'].value = f'Weekly Stats {date_string}'
        if stats_type == 'Monthly':
            sheet['N31'].value = f'Montly Stats {date_string}'
        if stats_type == 'Yearly':
            sheet['N31'].value = f'Yearly Stats {date_string}'

        sheet['N31'].alignment  = center_alignment 
        sheet['N31'].font = styles.Font(size=12, bold=True)
        sheet.column_dimensions['N'].width = 24.71
        sheet['P31'].value = 'Offered'
        sheet['P32'].value = '#'
        sheet['Q31'].value = 'Answered'
        sheet['Q32'].value = '#'
        sheet['R31'].value = 'Handle Time'
        sheet.column_dimensions['R'].width = 15
        sheet['R32'].value = 'Average'

        sheet.merge_cells('N33:O33')
        sheet.merge_cells('N34:O34')
        sheet.merge_cells('N35:O35')
        sheet.merge_cells('N36:O36')
        sheet.merge_cells('N37:O37')

        # Define the border style
        thin_border = styles.Border(
            left=styles.Side(style='thin'),
            right=styles.Side(style='thin'),
            top=styles.Side(style='thin'),
            bottom=styles.Side(style='thin')
        )

        # loop through each row in the range N33 to R37
        for row in sheet['N29:R37']:
            # loop through each cell in the row and apply the border
            for cell in row:
                cell.border = thin_border


        # add width to specific columns
        sheet.column_dimensions['A'].width = 49
        sheet.column_dimensions['D'].width = 14
        sheet.column_dimensions['K'].width = 15
        sheet.column_dimensions['O'].width = 10
        sheet.column_dimensions['Q'].width = 10

        # merge cells
        sheet.merge_cells('A1:A2')
        sheet.merge_cells('E1:F1')
        sheet.merge_cells('G1:H1')
        sheet.merge_cells('I1:J1')

        # Loop over all rows and columns and apply the centered alignment
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = center_alignment


        for row in range(31, 33):
            for col in ['N', 'O', 'P', 'Q', 'R']:
                cell = sheet[f'{col}{row}']
                cell.alignment = center_alignment
                cell.fill = blue
                cell.font = styles.Font(size=12, bold=True)
	

        start_row = 33
        for i, row in top_5.iterrows():
            sheet[f'N{start_row}'] = row['FullName']
            sheet[f'P{start_row}'] = row['calls_offered']
            sheet[f'Q{start_row}'] = row['calls_answered']
            sheet[f'R{start_row}'] = row['AHT']
            start_row += 1

        for row in range(33, start_row):
            for col in ['N', 'P', 'Q', 'R']:
                cell = sheet[f'{col}{row}']
                cell.border = thin_border
                cell.alignment = center_alignment
                cell.font = styles.Font(size=12)


        # background for aht in top 5
        for cell in sheet['R']:
            try:
                time_obj = datetime.strptime(str(cell.value), "%H:%M:%S").time()
                if isinstance(time_obj, time) and time_obj <= time(hour=0, minute=6, second=30):
                    cell.fill = green
                elif isinstance(time_obj, time) and time_obj >= time(hour=0, minute=6, second=31) and time_obj <= time(hour=0, minute=7, second=0):
                    cell.fill = orange
                else:
                    cell.fill = red
            except ValueError:
                pass

    
        # Save the workbook
    writer.close()
    os.startfile(output_path)


    # subject = None
    # recipient = 'lazlemlop@gmail.com'
    # body = None

    # # create the email subject dynamically
    # if stats_type == 'Daily':
    #     subject = f"{team_name} {stats_type} Stats {day_of_week} {date_string.replace('.', '/')}"
    # else:
    #     subject = f"{team_name} {stats_type} Stats {date_string.replace('.', '/')}"

    

    # send_email(subject, recipient, body)

# run_report_stats(date_string, file_path1)

