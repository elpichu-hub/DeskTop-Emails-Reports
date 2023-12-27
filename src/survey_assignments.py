import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import itertools
from datetime import datetime
import email_config
import os

# sets up email sending function to accept attachments
def send_email(subject, recipient, body, survey_file, surveySOP, cc=None, bcc=None):
    import email_config
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    import os
    from email.mime.base import MIMEBase
    from email import encoders
    

    EMAIL_ADDRESS = email_config.EMAIL_ADDRESS_AUTO
    EMAIL_PASSWORD = email_config.EMAIL_PASSWORD_AUTO

    # Create the email message
    message = MIMEMultipart()
    message['From'] = EMAIL_ADDRESS
    message['To'] = recipient
    message['Subject'] = subject
    message['Cc'] = cc
    message['Bcc'] = bcc

    message.attach(MIMEText(body, 'html'))

    if survey_file is not None:
        with open(survey_file, 'rb') as file:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(file.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', 'attachment', filename=os.path.basename(survey_file))
            message.attach(part)

    if surveySOP is not None:
        with open(surveySOP, 'rb') as file:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(file.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', 'attachment', filename=os.path.basename(surveySOP))
            message.attach(part)

    # Connect to the Gmail SMTP server and send the email
    with smtplib.SMTP('smtp.gmail.com', 587) as smtp:
        smtp.starttls()
        smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
        smtp.send_message(message)

# This will create the survey assignments and send the email
def survey_email_generator(file_path1, email_address, agents_for_survey, throughApp=False):

    df = pd.read_excel(file_path1)

    # Convert DataFrame to CSV
    csv_file = 'Survey Ticket Assignments.csv'
    df.to_csv(csv_file, index=False)

    # Predefined list of users
    users = agents_for_survey

    # Calculate the number of rows per user and the remainder
    num_rows = len(df)
    num_users = len(users)
    rows_per_user, extra_rows = divmod(num_rows, num_users)

    # Load the Excel file
    book = load_workbook(file_path1)

    # Select the first worksheet
    sheet = book.active

    # Mapping from hex color codes to color names
    color_names = {
        'FFC7CE': 'Pink',
        'FFEB9C': 'Yellow',
        'D9EAD3': 'LightGreen',
        'C9DAF8': 'LightBlue',
        'AB6C3E': 'Brown',
        'FDCFE8': 'Lavender',
        'FCE4D6': 'Peachpuff',
        'A2C4C9': 'Teal',
        'F4CCCC': 'Mistyrose',
        'DDEBF7': 'Blue',
        'EAD1DC': 'Mauve',
        'DD7E6B': 'Brick Red',
        'EA9999': 'Salmon',
        'B6D7A8': 'Sage',
        'A2C4C9': 'Teal',
        'AB6C3E': 'Brown',
        'FDCFE8': 'Lavender',
        'C6EFCE': 'Green',
    }

    colors = list(color_names.keys())

    # Create an iterable of users repeated
    user_cycle = itertools.cycle(users)

    # Reset fill for all cells
    empty_fill = PatternFill(fill_type=None)
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=9, max_row=num_rows + 1):  # Adjusting for Excel's 1-indexing
        for cell in row:
            cell.fill = empty_fill

    # Assign rows to each user and highlight them
    start_row = 2  # Assuming the data starts from the second row
    current_row = start_row

    user_colors = {}

    for i, user in enumerate(user_cycle):
        # Break the loop if we've assigned all rows
        if current_row > num_rows + start_row - 1:  # Adjusting for Excel's 1-indexing
            break

        # Calculate the number of rows for the current user
        # Give the extra rows to the first few users
        num_rows_for_this_user = rows_per_user + 1 if i < extra_rows else rows_per_user
        
        # Assign color to the user
        color = colors[i % len(colors)]
        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        user_colors[user] = color  # Store the color for later
        
        for _ in range(num_rows_for_this_user):
            # Apply color to the row and add the agent name in the next column
            for cell in sheet[current_row]:
                cell.fill = fill
            agent_cell = sheet.cell(row=current_row, column=9)  # Assuming the agent name should be in column I
            agent_cell.value = user

            current_row += 1

    # Adjust the width of column I
    sheet.column_dimensions['I'].width = 25

    # Get the current date and time
    now = datetime.now()

    # Format the date and time
    formatted_date = now.strftime("%A %m-%d-%y")

    # Save the modified Excel file
    modified_file = f'Survey Ticket Assignments {formatted_date}.xlsx'
    book.save(modified_file)

    # if throughApp == False: then means it's through email prepares the email
    if throughApp == False:
        surveySOP = 'surveySOP.png'

        greeting = ""

        # Determine the appropriate greeting based on the current time
        if now.hour < 12:
            greeting = "Good morning"
        elif now.hour < 18:
            greeting = "Good afternoon"
        else:
            greeting = "Good evening"

        subject = f'Survey Ticket Assignments {formatted_date}'
        # Add HTML tags
        body = "<html><body>"

        # Add result string and signature
        result_string = f"<h3>{greeting} Team, <br><br>We kindly request that you complete the assigned surveys by the end of the day. Once you have finished, please notify us through Teams. Your prompt response is greatly appreciated. Thank you for your dedication and efforts. The surveys can be found in the attached Excel file.<br><br> For customers who has scored a two or lower on the Survey and did not request a callback or provide a contact number, we ask that you send us their email information. Kindly note this in their account if possible. Please ensure to include all these customers' email addresses in a single email message to avoid clutter. </h3>"

        body += result_string.replace('\n', '<br>')

        # Add the color information for each agent
        color_info = ""
        for user, color in user_colors.items():
            color_name = color_names[color]
            line = f"<p style='margin: 5px;'><strong>{user} = {color_name}</strong></p>"
            color_info += line

        # Wrap the color information in a div with a max-width
        color_info = f"<div style='width: 600px;'>{color_info}</div>"

        # Add the color information to the email body
        body += color_info

        # Add signature to the body
        body += email_config.team_lead_elite_signature

        # Close HTML tags
        body += "</body></html>"

        try:
            send_email(subject=subject, recipient=email_address, body=body, survey_file=modified_file, surveySOP=surveySOP)
            print('send_email executed successfully')
        except Exception as e:
            print('send_email failed to execute')
            print(e)

    else:
        os.startfile(modified_file)
