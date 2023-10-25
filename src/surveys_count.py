import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, PatternFill, Font
import os
import email_config
from fuzzywuzzy import process

def preprocess_string(s):
    """Remove extra spaces and convert to lowercase."""
    return ' '.join(s.split()).lower()

# Read the CSV file named "TicketExport.csv" from the current directory
csv_file = 'TicketExport.csv'
df = pd.read_csv(csv_file)

# Preprocess the 'Ticketing Agent' names in the dataframe
df['Ticketing Agent'] = df['Ticketing Agent'].apply(preprocess_string)

# Count occurrences of each agent
agent_counts = df['Ticketing Agent'].value_counts()

# Create a DataFrame with Agent and Count columns
agent_df = pd.DataFrame({'Agent': agent_counts.index, 'Count': agent_counts.values})

# Create an Excel file with the Agent and Count columns
output_excel = 'AgentOccurrences.xlsx'
agent_df.to_excel(output_excel, index=False)

# Open the created Excel file and modify column widths and alignment
wb = Workbook()
ws = wb.active

# Load the existing data into the Excel sheet
for row in dataframe_to_rows(agent_df, index=False, header=True):
    ws.append(row)

# Set column widths and other formatting
ws.column_dimensions['A'].width = 26
ws.column_dimensions['B'].width = 10

for column in ws.columns:
    for cell in column:
        cell.alignment = Alignment(horizontal='center', vertical='center')

blue = PatternFill(start_color='0ACDFF', end_color='0ACDFF', fill_type='solid')
escalation_table = ws['I9']
escalation_table.value = "Escalation Agent"
escalation_table.alignment = Alignment(horizontal='center', vertical='center')
escalation_table.fill = blue
ws.column_dimensions['I'].width = 40

font = Font(size=14, bold=True)
escalation_table.font = font

escalation_table_total = ws['J9']
escalation_table_total.value = 'Total Surveys'
escalation_table_total.alignment = Alignment(horizontal='center', vertical='center')
escalation_table_total.fill = blue
ws.column_dimensions['J'].width = 20
escalation_table_total.font = font

# Preprocess the guardian agents list
guardians_agents = [preprocess_string(agent) for agent in email_config.guardians_agents]

def fuzzy_match(agent, agents_list):
    matches = process.extract(agent, agents_list, scorer=process.fuzz.partial_ratio)
    best_match = max(matches, key=lambda x: x[1])
    if best_match[1] >= 90:
        print(f"Matching {agent} to {best_match[0]} with score {best_match[1]}")
        return best_match[0]
    else:
        return None

# Step 1: Create the filtered_agents dictionary
filtered_agents_dict = {agent: fuzzy_match(agent, guardians_agents) for agent in agent_df['Agent']}

# Step 2: Create a new column 'MatchedAgent' in agent_df with the fuzzy-matched names
agent_df['MatchedAgent'] = agent_df['Agent'].map(filtered_agents_dict)

# Filter out rows where 'MatchedAgent' is not in guardians_agents
filtered_df = agent_df[agent_df['MatchedAgent'].isin(guardians_agents)].reset_index(drop=True).drop(columns=['MatchedAgent'])

print(filtered_df)

with pd.ExcelWriter(output_excel, engine='openpyxl', mode='a') as writer:
    filtered_df.to_excel(writer, sheet_name='Guardians Agents', index=False)

wb.save(output_excel)
print(f"Data saved to {output_excel}")

# Open the Excel file
os.system(f'start excel "{output_excel}"')
