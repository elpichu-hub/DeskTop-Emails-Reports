import os
import openpyxl
from openpyxl import styles
from openpyxl.styles import PatternFill
import datetime
from openpyxl.styles import Protection
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.protection import SheetProtection
import datetime
import email_function_with_attachments


# List of sheet names
sheet_names = ["Lazaro Gonzalez", "Dalila Ruiz", "Randall Lazaro", "Chiefy Chief", "Edgar Carneiro", "Karina Morales", "Lazaro Gonzalez", "Dalila Ruiz", "Randall Lazaro", "Chiefy Chief", "Edgar Carneiro", "Karina Morales"]
approved_codes = ['Close Account', 'Vehicle Not Found', 'Account Not Found', 'Credit Card Not Found', 'Credit Card Expired', 'Credit Card Declined', 'Credit Card Fraud']
# Rename the default first sheet to the first name in the list


# Define all the colors
blue = PatternFill(start_color='0ACDFF', end_color='0ACDFF', fill_type='solid')
grey = styles.PatternFill(start_color='808080', end_color='808080', fill_type='solid')
green = styles.PatternFill(start_color='59FFA0', end_color='59FFA0', fill_type='solid')
orange = styles.PatternFill(start_color='FF7F11', end_color='FF7F11', fill_type='solid')
red = styles.PatternFill(start_color='C1292E', end_color='C1292E', fill_type='solid')

# Define the dates
today = datetime.datetime.now()
tomorrow = datetime.datetime.now() + datetime.timedelta(days=1)
three_days_from_now = datetime.datetime.now() + datetime.timedelta(days=3)
# today = datetime.datetime(2024, 1, 1)
# tomorrow = today + datetime.timedelta(days=1)
# three_days_from_now = today + datetime.timedelta(days=3)
#### test link: https://conduent-my.sharepoint.com/:x:/r/personal/lazaro_gonzalez_conduent_com/_layouts/15/Doc.aspx?sourcedoc=%7B4658CE90-FA5E-4E74-93B1-1EC2DFF12FF3%7D&file=wrap_up_codes_12.29.2023.xlsx&action=default&mobileredirect=true&DefaultItemOpen=1&web=1



def protect_sheet(workbook, approved_codes):
    # Convert column letter to index
    col_index = column_index_from_string('O')

    # Define the range to allow editing
    min_row = 8
    max_row = 8 + len(approved_codes) - 1

    # Iterate through the sheets and protect/unprotect cells as needed
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Protect the entire sheet as read-only
        sheet.protection = SheetProtection(sheet=True)

        # Unprotect the specified range to allow editing
        for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=col_index, max_col=col_index):
            for cell in row:
                cell.protection = Protection(locked=False)


def create_wrap_up_codes_form(workbook, sheet_names, approved_codes):

    # Rename the default sheet to the first name in the list
    default_sheet = workbook.active
    default_sheet.title = sheet_names[0]

    # Rename the other sheets to the names in the list
    for sheet_name in sheet_names[1:]:
        workbook.create_sheet(sheet_name)

    # Loop through all sheets in the workbook
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Merge cells from column P to R in row 5
        sheet.merge_cells('M5:P5')
        for row in sheet['M5:P5']:
            for cell in row:
                cell.fill = red
                cell.font = styles.Font(bold=True, size=16)
                cell.alignment = styles.Alignment(horizontal='center', vertical='center')
                sheet.row_dimensions[5].height = 40
                cell.border = styles.Border(left=styles.Side(style='thin'), right=styles.Side(style='thin'), top=styles.Side(style='thin'), bottom=styles.Side(style='thin'))
                
        
        if today.weekday() == 0 or today.weekday() == 1 or today.weekday() == 2 or today.weekday() == 3 or today.weekday():
            sheet['M5'] = f'Wrap Up Codes {tomorrow.strftime("%m.%d.%Y")}'
        if today.weekday() == 4:
            sheet['M5'] = f'Wrap Up Codes {three_days_from_now.strftime("%m.%d.%Y")}'
        # Merge cells from column M to T in row 6
        sheet.merge_cells('M6:P6')

        # Apply the blue fill style to the merged cells
        for row in sheet['M6:P6']:
            for cell in row:
                cell.fill = blue
                cell.font = styles.Font(bold=False, size=14)
                cell.alignment = styles.Alignment(horizontal='center', vertical='center')
                sheet.row_dimensions[6].height = 40
                cell.border = styles.Border(left=styles.Side(style='thin'), right=styles.Side(style='thin'), top=styles.Side(style='thin'), bottom=styles.Side(style='thin'))

        sheet['M6'] = F'Wrap Up Codes {sheet_name}'
        
        # Merge cells from column M to N in row 7
        sheet.merge_cells('M7:N7')
        for row in sheet['M7:N7']:
            for cell in row:
                cell.fill = grey
                cell.font = styles.Font(bold=False, size=14)
                cell.alignment = styles.Alignment(horizontal='center', vertical='center')
                sheet.row_dimensions[7].height = 40
                cell.border = styles.Border(left=styles.Side(style='thin'), right=styles.Side(style='thin'), top=styles.Side(style='thin'), bottom=styles.Side(style='thin'))

        sheet['M7'] = 'Codes'

        # Merge cells from column O to P in row 7
        sheet.merge_cells('O7:P7')
        for row in sheet['O7:P7']:
            for cell in row:
                cell.fill = grey
                cell.font = styles.Font(bold=True)
                cell.alignment = styles.Alignment(horizontal='center', vertical='center')
                sheet.row_dimensions[7].height = 15
                cell.border = styles.Border(left=styles.Side(style='thin'), right=styles.Side(style='thin'), top=styles.Side(style='thin'), bottom=styles.Side(style='thin'))
        
        sheet['O7'] = 'Count'

        # Merge columns M and N for rows starting from row 8
        for row in range(8, 8 + len(approved_codes)):
            sheet.merge_cells(f'M{row}:N{row}')
            sheet.merge_cells(f'O{row}:P{row}')
            for row in sheet[f'M{row}:P{row}']:
                for cell in row:
                    cell.alignment = styles.Alignment(horizontal='center', vertical='center')
                    cell.border = styles.Border(left=styles.Side(style='thin'), right=styles.Side(style='thin'), top=styles.Side(style='thin'), bottom=styles.Side(style='thin'))
                    cell.font = styles.Font(size=14)

    # Iterate through the sheets and approved codes
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Convert the string into a list of characters
        codes = list(approved_codes)

        # Populate each character from codes in column M starting from row 8
        for i, code in enumerate(codes):
            row_number = 8 + i
            sheet[f'M{row_number}'] = code
            sheet.row_dimensions[row_number].height = 30

        sheet.column_dimensions['N'].width = 20

        # Add a total row after the last code
        total_row_number = 8 + len(codes)

        # Merge columns M and N for the total row
        sheet.merge_cells(f'M{total_row_number}:N{total_row_number}')
        sheet.merge_cells(f'O{total_row_number}:P{total_row_number}')

        total_count_cell = sheet[f'O{total_row_number}']
        total_count_cell.alignment = styles.Alignment(horizontal='center', vertical='center')
        total_count_cell.font = styles.Font(size=14)
        for col in range(column_index_from_string('O'), column_index_from_string('P') + 1):
            cell = sheet.cell(row=total_row_number, column=col)
            cell.border = styles.Border(left=styles.Side(style='thin'), 
                                        right=styles.Side(style='thin'), 
                                        top=styles.Side(style='thin'), 
                                        bottom=styles.Side(style='thin'))


        # Populate the total row
        total_cell = sheet[f'M{total_row_number}']
        total_cell.value = 'Total'
        total_cell.fill = green
        total_cell.font = styles.Font(bold=True, size=14)
        total_cell.alignment = styles.Alignment(horizontal='center', vertical='center')
        sheet.row_dimensions[total_row_number].height = 30

        # Assuming the counts are in column O, calculate the sum of counts
        sheet[f'O{total_row_number}'] = f'=SUM(O8:O{total_row_number - 1})'

        # Format the cells in column O for the total row
        total_count_cell = sheet[f'O{total_row_number}']
        # Apply border to the count cell in column O
        # Apply border to each cell in the merged range individually
        for col in range(column_index_from_string('M'), column_index_from_string('N') + 1):
            cell = sheet.cell(row=total_row_number, column=col)
            cell.border = styles.Border(left=styles.Side(style='thin'), 
                                        right=styles.Side(style='thin'), 
                                        top=styles.Side(style='thin'), 
                                        bottom=styles.Side(style='thin'))
            
        sheet.row_dimensions[7].height = 40
        sheet['O7'].font = styles.Font(bold=False, size=14)

    return 

def count_codes(approved_codes):
    workbook = None
    # Define a variable to hold the workbook filename
    workbook_filename = f'wrap_up_codes/wrap_up_codes_{today.strftime("%m.%d.%Y")}.xlsx'

    if today.weekday() in [0, 1, 2, 3, 4]:  # Weekdays
        print(f"Today is a weekday, and the date is {today.strftime('%m.%d.%Y')}")
        # Load the workbook
        workbook = openpyxl.load_workbook(workbook_filename)
        print(f"Workbook loaded from {workbook_filename}")

    counts = {code: 0 for code in approved_codes}
    agents_not_updated = []  # List to keep track of agents who haven't updated

    # Iterate through each sheet
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        total_count_for_sheet = 0  # Total count for each sheet

        # Iterate over each code and its corresponding count cell
        for i, code in enumerate(approved_codes):
            row_number = 8 + i
            count_cell = f'O{row_number}'  # Assuming count is in column O
            count = sheet[count_cell].value

            # Check if the count is a number and add it to the total
            if isinstance(count, (int, float)):
                counts[code] += count
                total_count_for_sheet += count

        # If total count for the sheet is 0, then add the agent to the list
        if total_count_for_sheet == 0:
            agents_not_updated.append(sheet_name)

    # Print the results
    for code, count in counts.items():
        print(f"Code '{code}' has a count of {count}")

    # Print agents who have not updated their wrap-up codes
    if agents_not_updated:
        print("Agents who have not updated their wrap-up codes:")
        email_function_with_attachments.send_email(subject=f"Wrap Up Codes Not Updated {today.strftime('%m.%d.%Y')}", recipient="lazarogonzalez.auto@gmail.com", body=f"Agents who have not updated their wrap-up codes: {agents_not_updated}", wrappup_file=workbook_filename)

        for agent in agents_not_updated:
            print(agent)
    else:
        print("All agents have updated their wrap-up codes.")
        email_function_with_attachments.send_email(subject=f"Wrap Up Codes Updated {today.strftime('%m.%d.%Y')}", recipient="lazarogonzalez.auto@gmail.com", body=f"All agents have updated their wrap-up codes.", wrappup_file=workbook_filename)


# Create a new workbook
workbook = openpyxl.Workbook()

# call functions
create_wrap_up_codes_form(workbook, sheet_names, approved_codes)
count_codes(approved_codes)
protect_sheet(workbook, approved_codes)







# Define the base OneDrive path (Windows style)
one_drive_base_path = r"C:\Users\etubr\OneDrive - Conduent\wrap_up_codes"

# Define the folder path within your OneDrive directory
folder_path = os.path.join(one_drive_base_path, "wrap_up_codes")

# Check if the folder exists, and create it if it doesn't
if not os.path.exists(folder_path):
    os.makedirs(folder_path)

# Determine the file name based on the day of the week
if today.weekday() in [0, 1, 2, 3]:
    file_name = f"wrap_up_codes_{tomorrow.strftime('%m.%d.%Y')}.xlsx"
elif today.weekday() == 4:
    file_name = f"wrap_up_codes_{three_days_from_now.strftime('%m.%d.%Y')}.xlsx"

# Create the full file path
file_path = os.path.join(folder_path, file_name)

# Save the workbook to the specified folder
workbook.save(file_path)

os.startfile(file_path)

# Print the saved file path
print(f"Workbook saved to {file_path}")